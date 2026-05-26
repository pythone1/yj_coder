"""
项目名称: aquaculture-census-platform
技术领域: 01-geospatial-remote-sensing
模块说明: gpkg.py - 核心业务算法实现
作者: 杨佳 (资深 AI 算法与遥感工程师)
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import geopandas as gpd
import pandas as pd
from openpyxl import load_workbook


def detect_header_row(excel_path: Path) -> int:
    wb = load_workbook(excel_path, read_only=True, data_only=False)
    ws = wb[wb.sheetnames[0]]
    row1 = [ws.cell(row=1, column=i).value for i in range(1, min(ws.max_column, 40) + 1)]
    row2 = [ws.cell(row=2, column=i).value for i in range(1, min(ws.max_column, 40) + 1)]
    wb.close()

    row1_text = {str(v).strip() for v in row1 if v is not None and str(v).strip()}
    row2_text = {str(v).strip() for v in row2 if v is not None and str(v).strip()}
    if {"基本情况", "养殖情况", "治理情况"} & row1_text and {"图斑编号", "养殖主体姓名（名称）"} <= row2_text:
        return 1
    return 0


def normalize_yes_no(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def classify_type(row: pd.Series) -> str:
    if normalize_yes_no(row["是否计划退养（包括转产）"]) == "是":
        return "退鱼转产"
    if normalize_yes_no(row["是否转养（调整养殖品种）"]) == "是":
        return "换鱼再养"
    if normalize_yes_no(row["是否计划空塘"]) == "是":
        return "空塘"
    if normalize_yes_no(row["是否为良种养殖"]) == "是":
        return "新鱼新养"
    return "老鱼优养"


def stringify_non_geometry_columns(df: pd.DataFrame) -> pd.DataFrame:
    converted = df.copy()
    for col in converted.columns:
        if col == "geometry":
            continue
        if pd.api.types.is_datetime64_any_dtype(converted[col]):
            converted[col] = converted[col].dt.strftime("%Y-%m-%d")
        elif converted[col].dtype == object:
            converted[col] = converted[col].apply(
                lambda x: x.strftime("%Y-%m-%d %H:%M:%S")
                if isinstance(x, datetime)
                else x
            )
    return converted


def main(
    excel_path: str | Path,
    gdb_path: str | Path,
    output_path: str | Path,
) -> None:
    excel_path = Path(excel_path)
    gdb_path = Path(gdb_path)
    output_path = Path(output_path)

    header_row = detect_header_row(excel_path)
    excel_df = pd.read_excel(excel_path, header=header_row)
    excel_df.columns = [str(c).strip() for c in excel_df.columns]
    excel_df = excel_df.loc[:, [c for c in excel_df.columns if c and not str(c).startswith("Unnamed:")]]

    required_cols = [
        "图斑编号",
        "是否计划退养（包括转产）",
        "是否转养（调整养殖品种）",
        "是否计划空塘",
        "是否为良种养殖",
    ]
    missing = [c for c in required_cols if c not in excel_df.columns]
    if missing:
        raise ValueError(f"Excel 缺少必要字段: {missing}")

    excel_df["图斑编号"] = excel_df["图斑编号"].astype(str).str.strip()
    excel_df["类型"] = excel_df.apply(classify_type, axis=1)

    gdf = gpd.read_file(gdb_path)
    if "tbid" not in gdf.columns:
        raise ValueError("图斑数据库缺少 tbid 字段")

    gdf["tbid"] = gdf["tbid"].astype(str).str.strip()
    matched_gdf = gdf[gdf["tbid"].isin(excel_df["图斑编号"])][["tbid", "geometry"]].copy()

    result = matched_gdf.merge(excel_df, left_on="tbid", right_on="图斑编号", how="left")
    result = result.drop(columns=["tbid"])
    result_gdf = gpd.GeoDataFrame(result, geometry="geometry", crs=gdf.crs)
    result_gdf = gpd.GeoDataFrame(stringify_non_geometry_columns(result_gdf), geometry="geometry", crs=gdf.crs)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.exists():
        output_path.unlink()
    result_gdf.to_file(output_path, driver="GPKG", layer="two_fish_info", encoding="utf-8")

    print(f"Excel 行数: {len(excel_df)}")
    print(f"匹配图斑数: {len(result_gdf)}")
    print(f"输出文件: {output_path}")
    print(result_gdf["类型"].value_counts(dropna=False).to_string())


if __name__ == "__main__":
    excel_path = r"E:\全省养殖池溏上图入库普查\PY\宜兴市\20251112\0319\0320两鱼信息表.xlsx"
    gdb_path = r"E:\全省养殖池溏上图入库普查\PY\宜兴市\20251112\0319\0320宜兴池塘图斑数据库"
    output_path = r"E:\全省养殖池溏上图入库普查\PY\宜兴市\20251112\0319\0320两鱼类型图斑.gpkg"

    main(excel_path=excel_path, gdb_path=gdb_path, output_path=output_path)
