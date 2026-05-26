from __future__ import annotations

import math
import re
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List

import geopandas as gpd
import pandas as pd
from openpyxl import load_workbook


TOTAL_AREA_HEADER = "养殖户总面积（亩）"
ALLOWED_SPECIES = {"鳊鱼、鲫鱼", "鳊鱼", "鲫鱼", "黄金鲫"}
YES_NO_COLUMNS = [
    "是否计划退养（包括转产）",
    "是否转养（调整养殖品种）",
    "是否计划空塘",
    "是否为良种养殖",
]
BASE_TARGET_HEADERS = [
    "序号",
    "图斑编号",
    "经纬度",
    "所在镇村（镇街+村）",
    "养殖主体姓名（名称）",
    "联系方式",
    "镇街定村（片）领导",
    "镇街监管员",
    "村定员干部",
    "村协管员",
    "塘口面积（亩）",
    "养殖户塘口数量（当前镇村）",
    "养殖品种",
    "池塘改造情况",
    "租金（元/亩）",
    "养殖密度（尾/亩）",
    "在塘水产品预计上市时间",
    "是否计划退养（包括转产）",
    "预计完成退养（转产）时间（到年月）",
    "是否转养（调整养殖品种）",
    "转养品种",
    "是否计划空塘",
    "2026年鳊鱼鱼种是否已落实",
    "鱼种来源及名称",
    "鱼种是否已放养",
    "是否为良种养殖",
    "备注",
]
REQUIRED_COLUMNS = {
    "图斑编号",
    "所在镇村（镇街+村）",
    "养殖主体姓名（名称）",
    "塘口面积（亩）",
    "养殖品种",
    "是否计划退养（包括转产）",
    "是否转养（调整养殖品种）",
    "是否计划空塘",
    "是否为良种养殖",
}
GROUP_KEY_COLUMNS = ["所在镇村（镇街+村）", "养殖主体姓名（名称）"]


@dataclass
class TemplateStyle:
    height: float | None
    styles: List[dict]


def build_target_headers(include_total_area: bool) -> List[str]:
    headers = BASE_TARGET_HEADERS.copy()
    if include_total_area:
        headers.insert(headers.index("养殖户塘口数量（当前镇村）"), TOTAL_AREA_HEADER)
    return headers


def normalize_text(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "nat"}:
        return ""
    return text


def trim_trailing_empty(values: List[object]) -> List[object]:
    items = list(values)
    while items and normalize_text(items[-1]) == "":
        items.pop()
    return items


def parse_area(value) -> float | None:
    if pd.isna(value):
        return None
    if isinstance(value, (int, float)) and not pd.isna(value):
        return float(value)
    text = normalize_text(value)
    if not text:
        return None
    cleaned = re.sub(r"[^\d.]+", "", text)
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def choose_group_area(values: Iterable[float | None]) -> float:
    numeric = [float(v) for v in values if v is not None and not math.isnan(v)]
    if not numeric:
        return 0.0
    counts = pd.Series(numeric).round(6).value_counts()
    top_count = counts.iloc[0]
    candidates = sorted(counts[counts == top_count].index.tolist())
    return float(candidates[-1])


def append_issue(issue_map: Dict[int, List[str]], row_index: int, message: str) -> None:
    issue_map.setdefault(row_index, [])
    if message not in issue_map[row_index]:
        issue_map[row_index].append(message)


def ensure_required_columns(df: pd.DataFrame) -> None:
    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(f"输入表缺少必要字段: {missing}")


def read_template_headers(template_path: Path) -> List[str]:
    wb = load_workbook(template_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(row=2, column=i).value for i in range(1, ws.max_column + 1)]
    wb.close()
    return trim_trailing_empty(headers)


def detect_input_header_row(input_path: Path) -> int:
    wb = load_workbook(input_path, read_only=True, data_only=False)
    ws = wb[wb.sheetnames[0]]
    first_row = [ws.cell(row=1, column=i).value for i in range(1, min(ws.max_column, 40) + 1)]
    second_row = [ws.cell(row=2, column=i).value for i in range(1, min(ws.max_column, 40) + 1)]
    wb.close()

    first_row_text = {normalize_text(v) for v in trim_trailing_empty(first_row) if normalize_text(v)}
    second_row_text = {normalize_text(v) for v in trim_trailing_empty(second_row) if normalize_text(v)}

    has_group_header = bool({"基础信息", "基本情况", "养殖情况", "治理情况"} & first_row_text)
    has_real_headers = {"图斑编号", "所在镇村（镇街+村）", "养殖主体姓名（名称）"} <= second_row_text
    return 1 if has_group_header and has_real_headers else 0


def capture_template_style(template_path: Path) -> TemplateStyle:
    wb = load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]
    source_row = 3 if ws.max_row >= 3 else 2
    styles = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=source_row, column=col_idx)
        styles.append(
            {
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
                "number_format": cell.number_format,
                "protection": copy(cell.protection),
                "_style": copy(cell._style),
            }
        )
    height = ws.row_dimensions[source_row].height
    wb.close()
    return TemplateStyle(height=height, styles=styles)


def load_input_dataframe(input_path: Path) -> pd.DataFrame:
    header_row = detect_input_header_row(input_path)
    df = pd.read_excel(input_path, header=header_row)
    df.columns = trim_trailing_empty([normalize_text(col) for col in df.columns])
    df = df.iloc[:, : len(df.columns)]
    if "鱼种否否已放养" in df.columns and "鱼种是否已放养" not in df.columns:
        df = df.rename(columns={"鱼种否否已放养": "鱼种是否已放养"})
    if "备注" not in df.columns:
        df["备注"] = ""
    return df


def prepare_base_columns(df: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()
    for col in result.columns:
        if result[col].dtype == object:
            result[col] = result[col].apply(lambda x: normalize_text(x) if isinstance(x, str) else x)
    result["图斑编号"] = result["图斑编号"].apply(normalize_text)
    result["所在镇村（镇街+村）"] = result["所在镇村（镇街+村）"].apply(normalize_text)
    result["养殖主体姓名（名称）"] = result["养殖主体姓名（名称）"].apply(normalize_text)
    result["_原始塘口面积"] = result["塘口面积（亩）"]
    result["_申报面积"] = result["塘口面积（亩）"].apply(parse_area)
    return result


def validate_and_enrich(
    df: pd.DataFrame,
    gdf: gpd.GeoDataFrame,
    include_total_area: bool,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    issue_map: Dict[int, List[str]] = {}
    result = df.copy()

    duplicated_ids = result["图斑编号"].duplicated(keep=False) & result["图斑编号"].ne("")
    for idx in result.index[duplicated_ids]:
        append_issue(issue_map, idx, "图斑编号在输入表中重复")

    species_series = result["养殖品种"].apply(normalize_text)
    invalid_species_mask = species_series.ne("") & ~species_series.isin(ALLOWED_SPECIES)
    for idx in result.index[invalid_species_mask]:
        append_issue(issue_map, idx, f"养殖品种不在允许范围内: {species_series.loc[idx]}")

    for col in YES_NO_COLUMNS:
        normalized = result[col].apply(normalize_text)
        invalid_mask = normalized.ne("") & ~normalized.isin({"是", "否"})
        for idx in result.index[invalid_mask]:
            append_issue(issue_map, idx, f"{col}只能填写“是”或“否”")

    transfer_yes = result["是否转养（调整养殖品种）"].apply(normalize_text).eq("是")
    transfer_species = result["转养品种"].apply(normalize_text)
    for idx in result.index[transfer_yes & transfer_species.isin({"", "/"})]:
        append_issue(issue_map, idx, "已填写转养，但“转养品种”为空")

    quit_yes = result["是否计划退养（包括转产）"].apply(normalize_text).eq("是")
    quit_time = result["预计完成退养（转产）时间（到年月）"].apply(normalize_text)
    quit_time = quit_time.where(quit_time.ne(""), other=result["预计完成退养（转产）时间（到年月）"])
    for idx in result.index[quit_yes & quit_time.apply(lambda x: normalize_text(x) in {"", "/"})]:
        append_issue(issue_map, idx, "已填写计划退养（转产），但“预计完成退养（转产）时间”为空")

    group_unique_areas = (
        result.groupby(GROUP_KEY_COLUMNS)["_申报面积"]
        .apply(lambda s: sorted({round(v, 6) for v in s.dropna().tolist()}))
        .reset_index(name="_面积集合")
    )
    inconsistent_groups = group_unique_areas[group_unique_areas["_面积集合"].apply(lambda x: len(x) > 1)]
    inconsistent_set = {
        (row["所在镇村（镇街+村）"], row["养殖主体姓名（名称）"]) for _, row in inconsistent_groups.iterrows()
    }
    for idx, row in result.iterrows():
        key = (row["所在镇村（镇街+村）"], row["养殖主体姓名（名称）"])
        if key in inconsistent_set:
            append_issue(issue_map, idx, "同镇同名分组内“塘口面积（亩）”填写不一致，已按组内众数/较大值分配")

    gdf = gdf.copy()
    gdf["tbid"] = gdf["tbid"].astype(str).str.strip()
    if gdf.crs is None:
        raise ValueError("池塘图斑库缺少坐标系信息，无法计算面积")
    gdf_projected = gdf.to_crs("EPSG:32650")
    gdf_projected["_图斑面积亩"] = gdf_projected.geometry.area * 0.0015
    area_df = pd.DataFrame(gdf_projected[["tbid", "_图斑面积亩"]])

    result = result.merge(area_df, left_on="图斑编号", right_on="tbid", how="left")
    result["_匹配成功"] = result["tbid"].notna()
    for idx in result.index[~result["_匹配成功"]]:
        append_issue(issue_map, idx, "图斑编号未在池塘数据库中匹配到")

    household_area_map = result.groupby(GROUP_KEY_COLUMNS)["_申报面积"].apply(choose_group_area).to_dict()
    result["_分组申报面积"] = result.apply(
        lambda row: household_area_map[(row["所在镇村（镇街+村）"], row["养殖主体姓名（名称）"])],
        axis=1,
    )
    if include_total_area:
        result[TOTAL_AREA_HEADER] = result["_分组申报面积"].round(2)

    result["养殖户塘口数量（当前镇村）"] = result.groupby(GROUP_KEY_COLUMNS)["图斑编号"].transform("count")
    result["_组内匹配图斑总面积"] = result.groupby(GROUP_KEY_COLUMNS)["_图斑面积亩"].transform("sum").fillna(0.0)
    result["塘口面积（亩）"] = 0.0

    for _, group in result.groupby(GROUP_KEY_COLUMNS, sort=False):
        group_idx = list(group.index)
        matched = group[group["_匹配成功"] & group["_图斑面积亩"].fillna(0).gt(0)]
        declared_area = float(group["_分组申报面积"].iloc[0])
        total_geom_area = float(matched["_图斑面积亩"].sum())

        if declared_area <= 0:
            for idx in group_idx:
                append_issue(issue_map, idx, "分组申报面积为空或无效，重分配后面积记为0")
            continue

        if total_geom_area <= 0:
            for idx in group_idx:
                append_issue(issue_map, idx, "该分组没有可用的匹配图斑面积，无法按遥感图斑分配")
            continue

        rounded_allocations: Dict[int, float] = {}
        for idx, row in matched.iterrows():
            rounded_allocations[idx] = round(declared_area * float(row["_图斑面积亩"]) / total_geom_area, 2)

        diff = round(declared_area - sum(rounded_allocations.values()), 2)
        if abs(diff) > 0 and rounded_allocations:
            adjust_idx = matched["_图斑面积亩"].idxmax()
            rounded_allocations[adjust_idx] = round(rounded_allocations[adjust_idx] + diff, 2)

        for idx in group_idx:
            result.at[idx, "塘口面积（亩）"] = rounded_allocations.get(idx, 0.0)

    normalized_remark = result["备注"].apply(normalize_text)
    issue_texts = []
    for idx in result.index:
        issue_list = issue_map.get(idx, [])
        existing = normalized_remark.loc[idx]
        merged_text = "；".join(([existing] if existing else []) + issue_list)
        issue_texts.append(merged_text)
    result["备注"] = issue_texts

    issue_rows = []
    for idx, row in result.iterrows():
        for message in issue_map.get(idx, []):
            issue_rows.append(
                {
                    "序号": idx + 1,
                    "图斑编号": row["图斑编号"],
                    "所在镇村（镇街+村）": row["所在镇村（镇街+村）"],
                    "养殖主体姓名（名称）": row["养殖主体姓名（名称）"],
                    "问题描述": message,
                }
            )
    issue_df = pd.DataFrame(issue_rows)

    summary_items = [
        ("输入记录数", len(result)),
        ("输入图斑编号唯一值数量", result["图斑编号"].nunique()),
        ("匹配成功记录数", int(result["_匹配成功"].sum())),
        ("未匹配记录数", int((~result["_匹配成功"]).sum())),
        ("存在问题的记录数", int(sum(bool(issue_map.get(idx)) for idx in result.index))),
        ("存在问题的分组数", int(len(inconsistent_set))),
    ]
    summary_df = pd.DataFrame(summary_items, columns=["指标", "值"])

    result["序号"] = range(1, len(result) + 1)
    return result, issue_df, summary_df


def to_excel_value(value):
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, datetime):
        return value
    if pd.isna(value):
        return None
    if isinstance(value, float):
        return float(value)
    return value


def apply_cell_style(target_cell, style_info: dict) -> None:
    target_cell._style = copy(style_info["_style"])
    target_cell.font = copy(style_info["font"])
    target_cell.fill = copy(style_info["fill"])
    target_cell.border = copy(style_info["border"])
    target_cell.alignment = copy(style_info["alignment"])
    target_cell.number_format = style_info["number_format"]
    target_cell.protection = copy(style_info["protection"])


def write_output(
    result_df: pd.DataFrame,
    issue_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    template_path: Path,
    output_path: Path,
    include_total_area: bool,
) -> None:
    style = capture_template_style(template_path)
    wb = load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]
    template_headers = trim_trailing_empty(
        [ws.cell(row=2, column=i).value for i in range(1, ws.max_column + 1)]
    )

    if include_total_area and TOTAL_AREA_HEADER not in template_headers:
        insert_at = build_target_headers(True).index(TOTAL_AREA_HEADER) + 1
        base_style_index = min(insert_at - 2, len(style.styles) - 1)
        base_style = style.styles[base_style_index]
        ws.insert_cols(insert_at)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=insert_at)
            apply_cell_style(cell, base_style)
            cell.value = TOTAL_AREA_HEADER if row_idx == 2 else None
        style.styles.insert(insert_at - 1, copy(base_style))

    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)

    output_headers = trim_trailing_empty(
        [ws.cell(row=2, column=i).value for i in range(1, ws.max_column + 1)]
    )
    for row_offset, (_, row) in enumerate(result_df.iterrows(), start=3):
        for col_idx, header in enumerate(output_headers, start=1):
            cell = ws.cell(row=row_offset, column=col_idx)
            cell.value = to_excel_value(row[header]) if header in result_df.columns else None
            apply_cell_style(cell, style.styles[min(col_idx - 1, len(style.styles) - 1)])
        if style.height is not None:
            ws.row_dimensions[row_offset].height = style.height

    if "校验明细" in wb.sheetnames:
        del wb["校验明细"]
    detail_ws = wb.create_sheet("校验明细")
    detail_headers = list(issue_df.columns) if not issue_df.empty else ["序号", "图斑编号", "所在镇村（镇街+村）", "养殖主体姓名（名称）", "问题描述"]
    detail_ws.append(detail_headers)
    if issue_df.empty:
        detail_ws.append(["", "", "", "", "未发现问题"])
    else:
        for _, detail_row in issue_df.iterrows():
            detail_ws.append([to_excel_value(detail_row[col]) if col in detail_row else None for col in detail_headers])

    if "处理汇总" in wb.sheetnames:
        del wb["处理汇总"]
    summary_ws = wb.create_sheet("处理汇总")
    summary_ws.append(list(summary_df.columns))
    for _, summary_row in summary_df.iterrows():
        summary_ws.append([summary_row["指标"], summary_row["值"]])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()


def main(
    input_path: str | Path,
    template_path: str | Path,
    output_path: str | Path,
    pond_db_path: str | Path,
    include_total_area: bool = True,
) -> None:
    input_path = Path(input_path)
    template_path = Path(template_path)
    pond_db_path = Path(pond_db_path)
    output_path = Path(output_path)

    template_headers = read_template_headers(template_path)
    allowed_headers = {
        tuple(BASE_TARGET_HEADERS),
        tuple(build_target_headers(True)),
    }
    if tuple(template_headers) not in allowed_headers:
        raise ValueError(f"模板第二行字段与脚本预期不一致，请先确认模板结构: {template_headers}")

    df = load_input_dataframe(input_path)
    ensure_required_columns(df)
    prepared_df = prepare_base_columns(df)
    gdf = gpd.read_file(pond_db_path)
    if "tbid" not in gdf.columns:
        raise ValueError("池塘图斑数据库缺少 tbid 字段，无法按图斑编号匹配")

    result_df, issue_df, summary_df = validate_and_enrich(prepared_df, gdf, include_total_area=include_total_area)
    target_headers = build_target_headers(include_total_area)
    for header in target_headers:
        if header not in result_df.columns:
            result_df[header] = ""
    result_df = result_df[target_headers]

    write_output(
        result_df=result_df,
        issue_df=issue_df,
        summary_df=summary_df,
        template_path=template_path,
        output_path=output_path,
        include_total_area=include_total_area,
    )

    print(f"输入文件: {input_path}")
    print(f"模板文件: {template_path}")
    print(f"池塘图斑库: {pond_db_path}")
    print(f"输出文件: {output_path}")
    print(f"是否计算养殖户总面积: {include_total_area}")
    print(f"处理记录数: {len(result_df)}")
    print(f"问题明细条数: {len(issue_df)}")


if __name__ == "__main__":
    input_path = r"E:\全省养殖池溏上图入库普查\PY\宜兴市\20251112\0319\徐舍修改.xlsx"
    template_path = r"E:\全省养殖池溏上图入库普查\PY\宜兴市\20251112\0319\导入信息表.xlsx"
    output_path = r"E:\全省养殖池溏上图入库普查\PY\宜兴市\20251112\0319\徐舍修改检查.xlsx"
    pond_db_path = r"E:\全省养殖池溏上图入库普查\养殖信息统计\20250902\宜兴市池塘图斑库.gpkg"
    include_total_area = True

    main(
        input_path=input_path,
        template_path=template_path,
        output_path=output_path,
        pond_db_path=pond_db_path,
        include_total_area=include_total_area,
    )
