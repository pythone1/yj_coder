"""
项目名称: aquaculture-census-platform
技术领域: 01-geospatial-remote-sensing
模块说明: ctxxtbyd.py - 核心业务算法实现
作者: 杨佳 (资深 AI 算法与遥感工程师)
"""

import os,glob

import geopandas as gpd
import pandas as pd
from shapely import wkt
from shapely.ops import unary_union


import numpy as np
import folium
import folium.plugins
import unicodedata
import re
# from LAC import LAC
import dimsim
from datetime import datetime, timedelta
import openpyxl

# lac = LAC(mode='lac')
SURNAME_PATH = r"E:\江苏省养殖池塘上图入库项目\进度统计表_原\行政区划\常见姓氏.txt"
XZQH = gpd.read_file(r'E:\python\省厅项目进度统计\JiangSu_XZQH.shp')
REGIONCODE = pd.read_excel(r'E:\江苏省养殖池塘上图入库项目\进度统计表_原\行政区划\行政编码.xlsx')

############################################################################
############################## 数据合并 #####################################
############################################################################
def mergeData(xls_file, polygon_file, dels_file=None):
    # 池塘轮廓
    st_time = datetime.now()
    ctlk = gpd.read_file(polygon_file)
    if 'id' in ctlk.columns:
        ctlk=ctlk.rename(columns={'id':'ID','tbid':'TBID'})
    ctlk = ctlk.rename(columns={'ID':'图斑id'})
    ctlk['图斑id'] = ctlk['图斑id'].astype('int').astype('str')
    ctlk = ctlk.set_index(['图斑id'], drop=False)
    ctlk['图斑面积'] = np.round(ctlk['area'] / 666.66, 2)
    ed_time = datetime.now()
    spd_time = (ed_time - st_time).total_seconds()
    print(f"池塘轮廓数据读取与处理：{np.round(spd_time/60,0)} 分钟")
    # ctlk.to_file(r'S:\项目数据\江苏省一池一档水产养殖基本情况普查项目\信息填报\tmp\ctlk.gpkg',encoding='utf-8',driver='GPKG')

    # 删除指定图斑
    if dels_file is not None:
        tbids = listDeletes(dels_file)
        ctlk = splitPolygons(ctlk,'TBID',tbids)
    else:
        ctlk['Ndel'] = True

    # 池塘信息
    st_time = datetime.now()
    if os.path.isdir(xls_file):
        files = glob.glob(f"{xls_file}\\*.xlsx")
        df_list = []
        for f in files:
            print(f"read {f}")
            df_list.append(pd.read_excel(f, dtype=str, skiprows=1))
        ctxx = pd.concat(df_list,ignore_index=True)
    else:
        ctxx = pd.read_excel(xls_file, dtype=str, skiprows=1)
        
    # 删除无效、审核驳回、未上报及测试数据 
    ctxx = deleteInvalidData(ctxx)
    ctxx.set_index('池塘id',inplace=True,drop=False)
    # 面积转数值型
    ctxx['合同面积'] = pd.to_numeric(ctxx['合同面积'],errors='coerce')
    ctxx['净水面面积'] = pd.to_numeric(ctxx['净水面面积'],errors='coerce') # 转数值型，非数值型强制转为nan
    # 校对状态添加数值型【状态值】
    ctxx = status2Num(ctxx)
    # # 养殖品种格式化
    # ctxx['养殖品种/预计亩产量'] = ctxx['养殖品种/预计亩产量'].str.replace('斤/亩','')
    # idx = ctxx[ctxx['养殖品种/预计亩产量']!='/'].index
    # for i in idx:
    #     yzxx = ctxx.loc[i,'养殖品种/预计亩产量'].split('，')
    #     pz = np.array([y.split(':')[0] for y in yzxx])
    #     cl = np.array([y.split(':')[1] for y in yzxx])
    #     sortedidx = np.argsort(pz)
    #     ctxx.loc[i,'养殖品种'] = ','.join(pz[sortedidx])
    #     ctxx.loc[i,'亩产量'] = ','.join(cl[sortedidx])
    # 转矢量
    ctxx['longitude'] = ctxx['池塘位置'].str.split('，', expand=True)[0]
    ctxx['latitude'] = ctxx['池塘位置'].str.split('，', expand=True)[1]
    ctxx = df2gdf(ctxx, 'longitude', 'latitude', epsg=4326).to_crs(ctlk.crs)
    ed_time = datetime.now()
    spd_time = (ed_time - st_time).total_seconds()
    print(f"池塘信息读取与处理：{np.round(spd_time/60,0)} 分钟")
    

    # 池塘信息合并TBID
    st_time = datetime.now()
    # ctxx = gpd.sjoin(ctxx, ctlk, how='left')
    try:
        ctxx = pd.merge(ctxx,ctlk.loc[:,['图斑id','TBID','图斑面积']],on=['图斑id'],how='left')
    except:
        ctxx = pd.merge(ctxx,ctlk.loc[:,['TBID','图斑面积']],on=['图斑id'],how='left')
    # ctxx.set_index('池塘id',inplace=True,drop=False)
    ed_time = datetime.now()
    spd_time = (ed_time - st_time).total_seconds()
    print(f"池塘信息合并轮廓ID：{np.round(spd_time/60,0)} 分钟")

    # 池塘信息填补合并编号
    ctxx.loc[ctxx['合并id'] == '/', '合并id'] = ctxx[ctxx['合并id'] == '/'].index
    # ctxx.to_file(r'S:\项目数据\江苏省一池一档水产养殖基本情况普查项目\信息填报\疑点和进度统计\总体进度\ctxx.gpkg',encoding='utf-8',driver='GPKG')

    # 排口位置
    st_time = datetime.now()
    pk = ctxx.copy()
    pk = pk[~pk['排口位置'].isnull()]
    pk = pk[pk['排口位置'] != '/']
    pk['longitude'] = pk['排口位置'].str.split('，', expand=True)[0]
    pk['latitude'] = pk['排口位置'].str.split('，', expand=True)[1]
    pk['wkt_str'] = 'POINT (' + pk['longitude'] + ' ' + pk['latitude'] + ')'
    pk['geometry'] = pk['wkt_str'].apply(wkt.loads)
    pk = pk.drop(columns=['longitude', 'latitude','wkt_str'])
    ed_time = datetime.now()
    spd_time = (ed_time - st_time).total_seconds()
    print(f"排口位置：{np.round(spd_time/60,0)} 分钟")


    # # 养殖品种重名的加水体类型为前缀重命名
    # npz = ['其他种类','南美白对虾','螺','鲈鱼']
    # for p in npz:
    #     idx0 = ctxx['养殖品种/预计亩产量'].str.contains(p)
    #     stlx = ctxx.loc[idx0,'水体类型'].unique()
    #     for s in stlx[stlx!='/']:
    #         idx = (ctxx['水体类型']==s) & (idx0)
    #         ctxx.loc[idx,'养殖品种/预计亩产量'] = ctxx.loc[idx,'养殖品种/预计亩产量'].str.replace(p,f"{s}{p}")

    # '''
    # 提取养殖品种
    # '''
    # ctxx= extractYZPZ(ctxx)

    # # 尾水检测位置
    # st_time = datetime.now()
    # ws = ctxx.copy()
    # ws = ws[~ws['检测塘口位置'].isnull()]
    # ws = ws[ws['检测塘口位置'] != '/']
    # ws['longitude'] = ws['检测塘口位置'].str.split('，', expand=True)[0]
    # ws['latitude'] = ws['检测塘口位置'].str.split('，', expand=True)[1]
    # ws['wkt_str'] = 'POINT (' + ws['longitude'] + ' ' + ws['latitude'] + ')'
    # ws['geometry'] = ws['wkt_str'].apply(wkt.loads)
    # ws = ws.drop(columns=['longitude', 'latitude','wkt_str'])
    # ed_time = datetime.now()
    # spd_time = (ed_time - st_time).total_seconds()
    # print(f"尾水检测位置：{np.round(spd_time/60,0)} 分钟")

    

    # 图斑增加填报信息
    st_time = datetime.now()
    ctlk = addAttr2Plygons(ctlk,ctxx)


    # # 养殖品种重名的加水体类型为前缀重命名
    # npz = ['其他种类','南美白对虾','螺','鲈鱼']
    # for p in npz:
    #     idx0 = ctlk['养殖品种/预计亩产量'].str.contains(p)
    #     stlx = ctlk.loc[idx0,'水体类型'].unique()
    #     for s in stlx[stlx!='/']:
    #         idx = (ctlk['水体类型']==s) & (idx0)
    #         ctlk.loc[idx,'养殖品种/预计亩产量'] = ctlk.loc[idx,'养殖品种/预计亩产量'].str.replace(p,f"{s}{p}")

    # '''
    # 提取养殖品种
    # '''
    # ctlk= extractYZPZ(ctlk)


    ed_time = datetime.now()
    spd_time = (ed_time - st_time).total_seconds()
    print(f"图斑增加填报信息：{np.round(spd_time/60,0)} 分钟")

    # 图斑增加填报状态：已填报养殖、已填报非养殖、未填报
    st_time = datetime.now()
    ctlk = polygonStatus(ctlk)
    ctxx = pointStatus(ctxx)
    ed_time = datetime.now()
    spd_time = (ed_time - st_time).total_seconds()
    print(f"图斑、填报点增加填报状态：{np.round(spd_time/60,0)} 分钟")

    # # 已填报图斑，按填报地址更新图斑所属区域——取消
    # idx = ctlk['填报状态'].str.contains('已填报')
    # dz = ctlk.loc[idx,'地址'].str.split('-',expand=True)
    # try:
    #     ctlk.loc[idx,'市'] = dz.loc[idx,1]
    #     ctlk.loc[idx,'区县'] = dz.loc[idx,2]
    # except:
    #     ctlk.loc[idx,'地方市'] = dz.loc[idx,1]
    #     ctlk.loc[idx,'地方区县'] = dz.loc[idx,2]

    return ctxx, ctlk, pk

def listDeletes(pth):
    '''
    汇总提报删除图斑
    '''
    df_list = []
    files = glob.glob(f"{pth}\\*.xlsx")
    for f in files:
        dfs = pd.read_excel(f,sheet_name=None)
        df_list.append(pd.concat(dfs.values(),ignore_index=True))
    df = pd.concat(df_list,ignore_index=True)
    df['TBID'] = df['TBID'].str[0:-5] + ',' + df['TBID'].str[-5:]

    return df['TBID'].unique()

def splitPolygons(df,field,values):
    '''
    根据字段值将图斑分为两个数据集，正常-上报删除
    '''
    if 'ID' not in df.columns:
        df['ID'] = df.index
    df0 = df.set_index(field,drop=False)
    df0['Ndel'] = True   # 用于进度统计
    
    vls0 = df[field].values
    idx = [True if v in vls0 else False for v in values]
    df0.loc[values[idx],'Ndel'] = False   # 上报删除
    df0 = df0.set_index('ID',drop=False) 

    return df0

def addAttr2Plygons(ctlk,ctxx):
    '''
    给图斑增加属性信息
    '''
    try:
        ctlk = pd.merge(ctlk,ctxx.loc[:, ['池塘id', '养殖经营人名称','养殖主体类型','状态','养殖品种/预计亩产量','水体类型','状态值','池塘所有权人名称','疑点信息','地址','图斑id']],on='图斑id',how='left')
    except:
        ctlk = pd.merge(ctlk.drop(columns='图斑id'),ctxx.loc[:, ['池塘id', '养殖经营人名称','养殖主体类型','状态','养殖品种/预计亩产量','水体类型','状态值','池塘所有权人名称','疑点信息','地址','图斑id']],on='图斑id',how='left')
    # ctlk = gpd.sjoin(ctlk, ctxx.loc[:, ['geometry', '养殖经营人名称','养殖主体类型','状态','状态值','池塘所有权人名称','疑点信息','地址','图斑id']],how='left')
    

    if 'ID' not in ctlk.columns:
        ctlk['ID'] = ctlk.index
    ctlk.reset_index(inplace=True,drop=True)
    ctlk.loc[ctlk['状态值'].isnull(),'状态值'] = -1 # 未填报，0为返回
    idx = ctlk.groupby('图斑id')['状态值'].idxmax().values
    ctlk = ctlk.loc[idx,:]
    ctlk = ctlk.set_index('ID',drop=False)

    return ctlk

def status2Num(df):
    '''
    将状态转为数值型
    '''
    df.loc[df['状态']=='待校对（村）','状态值'] = 1
    df.loc[df['状态']=='待校对（镇）','状态值'] = 2
    df.loc[df['状态']=='待校对（区县）','状态值'] = 3
    df.loc[df['状态']=='通过','状态值'] = 4
    df.loc[df['状态'].str.contains('返回'),'状态值'] = 0
    df.loc[df['状态值'].isnull(),'状态值'] = -1 # 未上报，应该在deleteInvalid已删除

    return df 


def df2gdf(df,lon_col,lat_col,epsg=4326):
    '''
    功能：df对象转gdf
    df: pd.DataFrame 数据表，含经纬度数据列
    lon_col: str 经度对应的列名
    lat_col: str 纬度对应的列名
    epsg: int 坐标系对应的编号，默认4326 WGS-1984
    '''
    df['wkt_str'] = 'POINT (' + df[lon_col] + ' ' + df[lat_col] + ')'
    
    df['geometry'] = df['wkt_str'].apply(wkt.loads)
    gdf = gpd.GeoDataFrame(df,crs='EPSG:'+str(epsg),geometry=df['geometry'])
    gdf = gdf.drop('wkt_str',axis=1)

    return gdf

def polygonStatus(gdf):
    '''
    图斑填报状态：已填报养殖、已填报非养殖、未填报
    '''
    naquadict = {
        '养殖经营人名称':['未养殖','退养','水库','光伏'],
        '池塘所有权人名称':['/']
    } # 待软件增加字段

    for k in list(naquadict.keys()):
        for v in naquadict[k]:
            gdf.loc[gdf[k]==v,'填报状态'] = '已填报非养殖'

    gdf.loc[gdf['填报状态'].isnull() & gdf['状态'].isnull(),'填报状态'] = '未填报' # 剩余填报养殖和未填报图斑中，未匹配到点（状态为空）的为未填报
    gdf.loc[gdf['填报状态'].isnull(),'填报状态'] = '已填报养殖'

    return gdf

def pointStatus(gdf):
    '''
    图斑填报状态：已填报养殖、已填报非养殖、未填报
    '''
    naquadict = {
        '养殖经营人名称':['未养殖','退养','水库','光伏'],
        '池塘所有权人名称':['/']
    } # 待软件增加字段

    for k in list(naquadict.keys()):
        for v in naquadict[k]:
            gdf.loc[gdf[k]==v,'填报状态'] = '已填报非养殖'

    gdf.loc[gdf['填报状态'].isnull(),'填报状态'] = '已填报养殖'
            
    return gdf


def deleteInvalidData(df):
    '''
    删除无效数据
    df: pd.DataFrame 填报数据表
    '''
    # 删除无效数据
    n1 = len(df)
    df = df[~df['状态'].isnull()]
    df = df[~df['池塘位置'].isnull()]
    df = df[df['池塘位置']!='/']
    df = df[df['池塘位置'].str.contains('，')]
    n2 = len(df)
    print(f"删除{n1-n2}个无效数据（错位）")

    # # 删除审核驳回、未上报
    # df = df[(~df['状态'].str.contains('未上报')) & (~df['状态'].str.contains('已返回'))]
    # 删除审核驳回、未上报
    df = df[~df['状态'].str.contains('未上报')]
    n3 = len(df)
    print(f"删除{n2-n3}个未上报数据")
 
    # 删除测试数据
    for c in ['养殖经营人名称','池塘所有权人名称']:
        # df = deleteTestData1(df, c, ['李四', '王五', '123'])  # 完全匹配删除
        df = deleteTestData2(df, c, ['测试'])  # 模糊匹配删除
    n4 = len(df)
    print(f"删除{n3-n4}个测试数据")
    
    return df

def deleteTestData1(df,field,keywords):
    '''
    按关键字keywords完全匹配删除测试数据
    '''
    # 删除测试数据
    values = df[field].values.tolist()

    test_idx = np.array([True if v in keywords else False for v in values])
    df = df[~test_idx]

    return df 

def deleteTestData2(df,field,keywords):
    '''
    按关键字keywords模糊匹配删除测试数据
    '''
    for k in keywords:
        df[field] = df[field].str.replace(k,'测试')

    # 删除测试数据
    values = df[field].values.tolist()
    
    test_idx = np.array([True if '测试' in v else False for v in values])
    df = df[~test_idx]

    return df 


############################################################################
############################## 疑点分析 #####################################
############################################################################

def pkAnalysis(sjoins,polygons,psxx):

    sjoins['疑点信息']=sjoins['疑点信息'].str.replace('排口位置较远','')
    sjoins['疑点信息']=sjoins['疑点信息'].str.replace('，排口位置较远','')
    sjoins['疑点信息']=sjoins['疑点信息'].str.replace('排口位置较远，','')

    sjoins['排口疑点'] = '无异常'
    sjoins = sjoins.to_crs('epsg:32650')
    polygons = polygons.to_crs('epsg:32650')
    psxx = psxx.to_crs('epsg:32650')

    sjoins_group = sjoins.groupby('合并id')
    sjoins_group_tbid = sjoins_group['图斑id'].unique()
    sjoins_group_ctid = sjoins_group['池塘id'].unique()
    sjoins_group_pknm = sjoins_group['排口位置'].unique().apply(lambda x: len(x))
    for i in sjoins_group_tbid.index:
        if sjoins_group_pknm[i] > 0:
            tbid = sjoins_group_tbid[i]
            ctid = sjoins_group_ctid[i]
            try:
                # tb = polygons.loc[tbid,['geometry']].dissolve().buffer(200).values[0]
                tb = polygons.loc[polygons[polygons['图斑id'].isin(tbid)].index,['geometry']].dissolve().buffer(200).values[0]
            except:
                # tb = sjoins.loc[ctid,['geometry']].dissolve().buffer(500).values[0]
                tb = sjoins.loc[sjoins[sjoins['池塘id'].isin(ctid)].index,['geometry']].dissolve().buffer(500).values[0]

        
            for j in ctid:
                if j in psxx['池塘id']:
                    pk = psxx.loc[j,'geometry']
                    if not pk.intersects(tb):
                        sjoins.loc[j,'排口疑点'] = '排口标记较远'


    return sjoins



def wsAnalysis(sjoins,polygons,psxx):

    # sjoins['疑点信息']=sjoins['疑点信息'].str.replace('排口位置较远','')
    # sjoins['疑点信息']=sjoins['疑点信息'].str.replace('，排口位置较远','')
    # sjoins['疑点信息']=sjoins['疑点信息'].str.replace('排口位置较远，','')

    sjoins['尾水检测位置疑点'] = '无异常'
    sjoins = sjoins.to_crs('epsg:32650')
    polygons = polygons.to_crs('epsg:32650')
    psxx = psxx.to_crs('epsg:32650')



    for i in psxx.index:
        tbid=psxx.loc[i,['图斑id']]
        # print(len(tbid))
        if tbid.values[0]!='/':
            sjoins.loc[i,'尾水检测位置疑点'] = '无对应图斑'
        else:
            tb = polygons.loc[polygons[polygons['图斑id'].isin(tbid)].index,['geometry']].values[0]
            wsxx = psxx.loc[i,'geometry']
            if not wsxx.intersects(tb):
                sjoins.loc[i,'尾水检测位置疑点'] = '尾水检测点在池塘外'
  



    return sjoins


def AQUATPYDAnalysis(sjoins):
    sjoins['养殖方式疑点'] = '无异常'
    idx1 = sjoins['养殖方式'] == '跑道养鱼'
    idx2 = np.array([True if '蟹' in sjoins['养殖品种/预计亩产量'].values else False])
    idx3 = np.array([True if '虾' in sjoins['养殖品种/预计亩产量'].values else False])
    idx = (idx1 & idx2) | (idx1 & idx3)
    sjoins.loc[idx,'养殖方式疑点'] = '养殖方式与养殖名称冲突'

    return sjoins


def is_chinese(str_char):
    str_char = str_char.strip()
    for char in str_char:
        if 'CJK' not in unicodedata.name(char):
            return False
    return True

def is_person(str_char):
    str_char = str_char.strip()
    if not is_chinese(str_char):
        return False
    if len(str_char) > 4:
        return False
    
    pattern = ['村', '街', '社区', '集体', '殖', '养殖场', '农场', '钓场', 
               '公司','居委会','委员会','老百姓','九组','八组',
               '七组','六组','五组','四组','三组','二组','一组','十一',
               '种牛场','种蜂场','林场']      
    for p in pattern:
        if p in str_char:
            return False
    
    
    # with open(SURNAME_PATH, 'r', encoding='UTF-8') as f:
    #     data = f.read()
    #     f.close()
    # p_surname = data.split(',')
    # for p in p_surname:
    #     if p == str_char[:len(p)]:
    #         return True
    return True

# 判断字符串是否为集体/公司
def is_company(str_char):
    str_char = str_char.strip()
    
    pattern1 = [r'(', r')', r'（', r'）']    
    for p in pattern1:
        str_char = str_char.replace(p, r'')
    
    if not is_chinese(str_char):
        return False
    
    pattern2 = ['镇', '村', '街', '社区', '集体', '居委会', '组','农庄','基地',
                '合作社', '养殖场', '农场', '钓场','生产队','处理中心','发展中心','养殖中心','水务站',
                '厂', '公司', '集团', '研究院','局','种牛场','部','种蜂场','闸管所','高级中学',
                '禽业','能源','委员会','生态园','林场', '农科院','电站','育苗场','管理所','病院',
                '湾景区','农业园区','医院','牧业','风景区','管理所','促进中心','帕蒂亚庄园','猪场','张湾中学',
                '人民政府','指导站','经济开发区','张湾乡','闸坝所','管理处']
    for p in pattern2:
        if p in str_char:
            return True
   
    return False

# 判断人名是否为笔误（同读音）
def is_person_similar(name1, name2):
    threshold = 0.1
    pattern = [r' ', r' ']
    for p in pattern:
        name1 = name1.replace(p, r'')
        name2 = name2.replace(p, r'')
    if (len(name1) == len(name2)) and (is_person(name1)) and (is_person(name2)):
        if dimsim.get_distance(name1, name2) < threshold:
            return True
    else:
        return False

def isPhoneNumber(phonenumber):
    if (len(phonenumber) != 11) or (not phonenumber.isdigit()):
        return False
    return True

def contractPeriod(sjoins):
    '''
    承包期限疑点
    '''
    sjoins['CBQXYD'] = '无异常'
    time_difference = 0
    if time_difference.days < 10*365:
        return False
    
# def idnumberAndZTMCAnalysis(sjoins):
#     '''
#     分析同主体名、不同身份证号的疑点
#     '''
#     sjoins['身份证号疑点'] = ''
#     grouped = sjoins.groupby(['养殖经营人名称','身份证号'])
#     groups = grouped.size().index.values
#     name = [g[0] for g in groups]
#     idnumber = [g[1] for g in groups]
#     # 同证件号不同人名
#     dulplicates = getDulplicates(idnumber)
#     for d in dulplicates:
#         if d != '/':
#             idx = (sjoins['身份证号'] == d)
#             sjoins.loc[idx,'身份证号疑点'] += '同身份证号不同人名,'
#     # 同人名不同证件号
#     dulplicates = getDulplicates(name)
#     for d in dulplicates:
#         idx = (sjoins['养殖经营人名称'] == d)
#         sjoins.loc[idx,'身份证号疑点'] += '同人名不同身份证号,'

#     sjoins.loc[sjoins['身份证号疑点']=='','身份证号疑点'] = '无异常'

#     return sjoins

def idnumberAndZTMCAnalysis(sjoins):
    '''
    分析同主体名、不同身份证号的疑点
    '''
    sjoins['身份证号疑点'] = ''
    grouped = sjoins.groupby(['养殖经营人名称','身份证号'])
    groups = grouped.size().index.values
    idnumber = [g[1] for g in groups]
    # 同证件号不同人名
    dulplicates = getDulplicates(idnumber)
    dulplicates = dulplicates[dulplicates!='/']
    for d in dulplicates:
        idx = (sjoins['身份证号'] == d)
        sjoins.loc[idx,'身份证号疑点'] += '同身份证号不同人名,'
    # 同人名不同证件号
    grouped = sjoins.groupby(['地址','养殖经营人名称','身份证号'])
    groups = grouped.size().index.values
    name = ['-'.join(g[0:2]) for g in groups]
    dulplicates = getDulplicates(name)
    name_series = sjoins['地址'] + '-' + sjoins['养殖经营人名称']
    for d in dulplicates:
        idx = (name_series == d)
        sjoins.loc[idx,'身份证号疑点'] += '同人名不同身份证号,'

    sjoins.loc[sjoins['身份证号疑点']=='','身份证号疑点'] = '无异常'

    return sjoins

def MCAnalysis(df):
    for i in df.index:
        nature = df.loc[i, '池塘所有权']
        name = df.loc[i, '池塘所有权人名称']
        if (nature != '其他') & ('家庭农场' in name):
            yd = ''
        else:
            # 所有权为个人，所有权人名称不是人名
            if (nature == '个人') & (not is_person(name)):
                yd = '所有权人名称疑似填写有误（非人名）,'
            
            # 所有权为集体，所有权人名称不是集体/公司
            elif (nature == '集体公司') & (not is_company(name)):
                yd = '所有权人名称疑似填写有误（非集体/公司）,'
                    
            # 所有权为国有（其他），所有权名称不是人名且不是集体/公司
            elif (nature == '其他') & ((is_company(name) or is_person(name)) or ('家庭农场' in name)):
                yd = '所有权人名称疑似填写有误（集体/公司、个人）,' 

            else:
                yd = ''

        nature = df.loc[i, '养殖主体类型']
        name = df.loc[i, '养殖经营人名称']
        if (nature != '其他') & ('家庭农场' in name):
            yd += ''
        else:
            # 主体性质为个人，主体名称不是人名
            if (nature == '个人') & (not is_person(name)):
                yd += '主体名称疑似填写有误（非人名）,'
                    
            # 主体性质为集体/公司，但名称不是公司
            elif (nature == '集体/公司') & (not is_company(name)):
                yd += '主体名称疑似填写有误（非集体/公司）,'
                    
            # 主体性质为其他，但主体名称是个人或集体/公司
            elif (nature == '其他'):
                if (not is_chinese(name)):
                    yd += '主体名称疑似填写有误（非中文）,'
                
                if is_company(name) or is_person(name) or ('家庭农场' in name):
                    yd += '主体性质疑似填写有误,'        
            
            # 联系人不是人名
            name = df.loc[i, '联系人']
            if not is_person(name):
                yd += '联系人疑似填写有误（非人名）,'

            # 人名疑似笔误
            if is_person_similar(df.loc[i, '池塘所有权人名称'], df.loc[i, '养殖经营人名称']) and (df.loc[i, '池塘所有权人名称'] != df.loc[i, '养殖经营人名称']):
                yd += '人名疑似笔误,'

        df.loc[i,'名称疑点'] = yd[0:-1]

    df.loc[df['名称疑点']=='','名称疑点'] = '无异常'

    return df

def contractTimeAnalysis(sjoins):
    '''
    承包期限
    '''
    sjoins['承包期限疑点'] = ''
    #  所有权人名称和主体名称不一致，承包期限未填写
    idx = (sjoins['池塘所有权人名称']!=sjoins['养殖经营人名称']) & (sjoins['承包期限'].isnull() | sjoins['承包期限']=='/')
    sjoins.loc[idx,'承包期限疑点'] = '主体与所有权人不同但承包期限未填写'

    # 承包结束时间疑似过短
    # idx = (~sjoins['承包期限'].isnull()) & (sjoins['承包期限']!='/')
    # idx = sjoins[idx].index
    # for i in idx:
    #     contract_endtime = sjoins.loc[i, '承包期限'].split(' : ')[1]
    #     if is_person_similar(sjoins.loc[i, '池塘所有权人名称'], sjoins.loc[i, '养殖经营人名称']):
    #         time_difference = datetime.strptime(contract_endtime, "%Y-%m-%d") - datetime.now()
    #         if time_difference.days < 10*365:
    #             sjoins.loc[i, '承包期限疑点'] = sjoins.loc[i, '承包期限疑点'] + ',承包结束时间疑似过短' 

    # 养殖主体和所有权人一致，无承包期限
    idx = (sjoins['池塘所有权人名称']==sjoins['养殖经营人名称']) & ((~sjoins['承包期限'].isnull()) & (sjoins['承包期限']!='/'))
    sjoins.loc[idx,'承包期限疑点'] = '主体与所有权人相同应无承包期限'

    sjoins.loc[sjoins['承包期限疑点']=='','承包期限疑点'] = '无异常'

    return sjoins


def yieldAnayesis(df):
    '''
    产量疑点：单养河蟹亩产量小于100或大于500作为疑点
    '''
    df['亩产量疑点'] = '无异常'
    idx0 = df['养殖品种'] == '河蟹'
    subsets = df.loc[idx0,:]
    subsets['产量'] = subsets['产量'].astype(int)
    idx = subsets['产量'] < 100
    subsets.loc[idx,'亩产量疑点'] = '河蟹亩产量小于100'
    idx = subsets['产量'] > 500
    subsets.loc[idx,'亩产量疑点'] = '河蟹亩产量大于500'
    df.loc[idx0,'亩产量疑点'] = subsets['亩产量疑点']

    return df

def inXZQH(gdf1,gdf2,shi,xian):
    '''
    填报点是否在行政区划范围内
    '''
    # idx = (gdf2['市'] == shi) & (gdf2['区县'] == xian)
    idx = (gdf2['市'] == shi)
    gdf2 = gdf2[idx].to_crs('epsg:32650')
    gdf2.geometry = gdf2.geometry.buffer(100)
    gdf2 = gdf2.to_crs(gdf1.crs)
    idx = gdf1.geometry.intersects(unary_union(gdf2['geometry']))
    # idx = gdf1.geometry.intersects(gdf2.geometry.values[0])
    gdf1.loc[idx,'区划外疑点'] = '无异常'
    gdf1.loc[~idx,'区划外疑点'] = '在区划范围外'

    return gdf1

def locationAnalysis(sjoins, polygons):
    # 点不在图斑内
    sjoins['位置疑点'] = '无异常'
    # sjoins.loc[sjoins['index_right'].isnull(), '位置疑点'] = '无对应图斑'
    sjoins.loc[sjoins['图斑id']=='/', '位置疑点'] = '无对应图斑'

    # 1个图斑有多个点
    dulplicates = getDulplicates(sjoins.loc[sjoins['图斑id']!='/','图斑id'].values)  # ID来自polygon ID
    for d in dulplicates:
        idx = sjoins['图斑id'] == d
        idx_str = '、'.join(sjoins.loc[idx,'池塘id'].values.tolist())
        sjoins.loc[idx, '位置疑点'] = idx_str + '对应同一图斑'

    # 图斑内无点
    polygons['位置疑点'] = ''
    # registered = sjoins.loc[~sjoins['图斑id'].str.contains('/'),'图斑id'].values
    registered = sjoins.loc[sjoins['图斑id']!='/','图斑id'].unique()
    # polygons.loc[registered, '位置疑点'] = '已填报'
    polygons.loc[polygons[polygons['图斑id'].isin(registered)].index, '位置疑点'] = '已填报'
    polygons.loc[polygons['位置疑点'] == '', '位置疑点'] = '未填报'
    # polygons.loc[dulplicates, '位置疑点'] = '多次填报'
    polygons.loc[polygons[polygons['图斑id'].isin(dulplicates)].index, '位置疑点'] = '多次填报'

    

    return sjoins, polygons


def YZLXAnalysis(sjoins,lctable):
    sjoins['YZLXYD'] = '无异常'
    for i in sjoins.index:
        # 用户填报类型
        yzlx1 = sjoins.loc[i,'variety_name']
        if isinstance(yzlx1,str):
            try:
                yzlx1 = np.array([lctable.loc[y,'YZLX'] for y in yzlx1.split(',')])
                yzlx1 = np.nanmin(yzlx1)
                # 卫星监测类型
                if (~np.isnan(yzlx1)) & (sjoins.loc[i,'area']>5):
                    yzlx2 = sjoins.loc[i,'YZLX']
                    if yzlx2!=yzlx1:
                        sjoins.loc[i,'YZLXYD'] = '养殖类型疑点'
            except:
                continue

    return sjoins

def getDulplicates(data):
    '''
    获取有重复项的值
    '''
    a,b = np.unique(data,return_counts=True)
    
    return a[b>1]

def areaAnalysis1(gdf):
    '''
    分析面积疑点:矢量面积-水域面积
    '''
    areas1 = gdf['图斑面积'].values
    areas2 = gdf['净水面面积'].values
    diff = np.abs(areas2 - areas1)
    pct = diff / areas1
    results = np.array(['无异常']*len(diff),dtype='<U20')
    results[(areas1<=5) & (diff>1)] = '水面面积偏差大于1亩'
    results[(areas1>5) & (pct>0.5)] = '水面面积偏差大于50%'
    results[np.isnan(areas2)] = '无异常'
    gdf['水面面积疑点'] = results

    return gdf

# def areaAnalysis2(gdf):
#     '''
#     原版本：按合并编号统计填报面积唯一值与图斑总面积相比；若填报面积不唯一，报同一合同不同面积；若合同id查找不全，报有返回
#     分析面积疑点:矢量面积-合同面积
#     gdf:用户填报点 拼接 分割面信息
#     '''
#     gdf['合同面积疑点'] = ''
#     groupids = gdf['合并id'].astype('str').values
#     for g in groupids:
#         idx = g.split('、')
#         try:
#             area1 = gdf.loc[idx,'合同面积'].values
#             area1 = area1[~np.isnan(area1)]
#             area2 = gdf.loc[idx,'图斑面积'].values
#             area2 = np.sum(area2)
#             if len(np.unique(area1)) > 1:
#                 gdf.loc[idx,'合同面积疑点'] = '同一合同不同面积'
#             elif len(np.unique(area1)) == 0:
#                 gdf.loc[idx,'合同面积疑点'] = '无异常'
#             else:
#                 diff = np.abs(area1[0] - area2)
#                 pct = diff / area2
#                 if (area2<=5) & (diff>1):
#                     gdf.loc[idx,'合同面积疑点'] = '合同面积偏差大于1亩'
#                 elif (area2>5) & (pct>0.5):
#                     gdf.loc[idx,'合同面积疑点'] = '合同面积偏差大于50%'
#                 else:
#                     gdf.loc[idx,'合同面积疑点'] = '无异常'
#         except:
#             new_idx = []
#             for i in idx:
#                 if i in gdf.index:
#                     new_idx.append(i)
#             gdf.loc[new_idx,'合同面积疑点'] = '合并池塘ID中有池塘被返回'
    
#     return gdf

def areaAnalysis2(gdf):
    '''
    20250220新版本：查看是不是同一个合并编号，是的话，
    1、比较合同面积是否一致；
    2、如果合同面积一致，分别比较各池塘遥感图斑面积和合同面积、各池塘遥感图斑面积之和和合同面积，有一个一致就行；
    3、如果合同面积/净水面面积不一致，各池塘遥感面积分别比较各自填报的合同面积或净水面面积

    gdf:用户填报点 拼接 分割面信息
    '''
    gdf['合同面积疑点'] = ''

    # 计算总面积
    ht_group = gdf.groupby('合并id')
    htzmj = ht_group['合同面积'].unique().apply(lambda x: sum(x[~np.isnan(x)]))
    htmj_num = ht_group['合同面积'].unique().apply(lambda x: len(x[~np.isnan(x)])) # 非空、唯一值个数
    tbzmj = ht_group['图斑面积'].sum()
    for i in htzmj.index:
        gdf.loc[gdf['合并id']==i,['合同面积唯一值个数','合同总面积','图斑总面积']] = [htmj_num[i],htzmj[i],tbzmj[i]]
    
    areas1 = gdf['图斑面积'].values
    areas2 = gdf['合同面积'].values
    areas3 = gdf['图斑总面积'].values
    areas4 = gdf['合同总面积'].values
    diff21 = np.abs(areas2 - areas1)    
    pct21 = diff21 / areas1
    diff43 = np.abs(areas4 - areas3)
    pct43 = diff43 / areas3
    yd1 = ((areas1<=5) & (diff21>1)) | ((areas1>5) & (pct21>0.5)) # 单个比存在异常的记录，有问题为True
    yd2 = ((areas3<=5) & (diff43>1)) | ((areas3>5) & (pct43>0.5)) # 合并比存在异常的记录，有问题为True

    idx = (gdf['合同面积唯一值个数']==1) & yd1 & yd2
    gdf.loc[idx,'合同面积疑点'] = '合同面积偏差较大'

    idx = (gdf['合同面积唯一值个数']>1) & yd1
    gdf.loc[idx,'合同面积疑点'] = '合同面积偏差较大'

    gdf.loc[gdf['合同面积疑点']=='','合同面积疑点'] = '无异常'

    return gdf    

def areaAnalysis3(sjoins):
    '''
    分析合同面积是否有未合并
    '''
    sjoins['合并疑点'] = '无异常'
    grouped = sjoins[sjoins['身份证号']!='/'].groupby(['合同面积','身份证号'])
    group_idx = grouped.size().index
    for gi in group_idx:
        merge_pond_ids = grouped.get_group(gi)['合并id'].values
        # 同合同面积、身份证号，不同合并池塘ID
        # if len(np.unique(merge_pond_ids))>1:
        if len(np.unique(merge_pond_ids.tolist()))>1:
            area2 = grouped.get_group(gi)['图斑面积'].sum()
            area1 = gi[0]
            diff = np.abs(area1 - area2)
            pct = diff / area2
            idx = grouped.get_group(gi).index
            prefix = ','.join([str(i) for i in idx])
            if (area2>5) & (pct>0.5):
                sjoins.loc[idx,'合并疑点'] = prefix + '疑似忘记合并且面积偏差大于50%'
            elif (area2<=5) & (diff>1):
                sjoins.loc[idx,'合并疑点'] = prefix + '疑似忘记合并且面积偏差大于1亩'
            else:
                sjoins.loc[idx,'合并疑点'] = prefix + '疑似忘记合并'

    return sjoins


############################################################################
############################## 进度统计 #####################################
############################################################################

def TBJDTJ1(shpfile,qxfile,names=None):
    '''
    填报进度统计1：基于行政区划
    '''
    # shpfile = r'D:\项目数据\江苏省\疑点核查\宜兴市\疑点分析结果2.gpkg'
    # qxfile = r'D:\项目数据\江苏省\行政界限\行政界限.shp'
    # names = ['六圩村','志泉','南庄','南塘村','路庄村'] # qx['cun_name']

    ct = gpd.read_file(shpfile)
    qx = gpd.read_file(qxfile).to_crs(ct.crs)

    if names is None:
        names = qx['cun_name'].values.tolist()
    
    num0 = []
    areas0 = []
    num1 = []
    areas1 = []
    num2 = []
    areas2 = []
    for n in names:
        roi = qx.loc[qx['cun_name']==n,'geometry']
        ct_roi0 = ct[ct.geometry.intersects(roi.geometry.values[0])]
        # 总图斑
        num0.append(len(ct_roi0))
        areas0.append(np.round(ct_roi0['area'].sum()/666.6666))
        # 未填报图斑
        ct_roi1 = ct_roi0[ct_roi0['IDYD']=='未填报']
        num1.append(len(ct_roi1))
        areas1.append(np.round(ct_roi1['area'].sum()/666.6666))
        # 已填报图斑
        ct_roi2 = ct_roi0[ct_roi0['IDYD']!='未填报']
        num2.append(len(ct_roi2))
        areas2.append(np.round(ct_roi2['area'].sum()/666.6666))

    df = pd.DataFrame({'名称':names,
                    '总图斑数':num0,
                    '总图斑面积':areas0,
                    '已填报图斑数':num2,
                    '已填报图斑面积':areas2,
                    '未填报图斑数':num1,
                    '未填报图斑面积':areas1,})
    df.to_excel(shpfile.replace('.gpkg','_填报统计.xlsx'))
    

def TBJDTJ2(xls):
    '''
    填报进度统计2：基于用户填报详细地址分类
    '''
    if isinstance(xls,str):
        df = pd.read_excel(xls,index_col='id')
    else:
        df = xls

    xxdz = np.unique(df['地址'].values)

    cun_num = []
    cun_area = []
    zhen = []

    for k in xxdz:
        cun_num.append(len(df[df['地址']==k]))
        cun_area.append(np.round(df.loc[df['地址']==k,'图斑面积'].sum(),2))
        zhen.append('-'.join(k.split('-')[0:4]))

    cun_df = pd.DataFrame({
        '村/集体':xxdz,
        '已填报信息条数':cun_num,
        '已填报图斑面积(单位：亩)':cun_area,
        '镇':zhen
        })
    
    zhen_num = []
    zhen_area = []
    zhen = np.unique(cun_df['镇'].values)
    for k in zhen:
        zhen_num.append(cun_df.loc[cun_df['镇']==k,'已填报信息条数'].sum())
        zhen_area.append(cun_df.loc[cun_df['镇']==k,'已填报图斑面积(单位：亩)'].sum())

    zhen_df = pd.DataFrame({
        '镇':zhen,
        '已填报信息条数':zhen_num,
        '已填报图斑面积(单位：亩)':zhen_area,
        })

    return cun_df.drop(columns=['镇']),zhen_df

def TBJDTJ3(gdf,roi=None):
    ''' 
    按池塘图斑gdf统计某区县总体及各分区roi的填报进度 
    gdf: gpd.GeoDataFrame，池塘图斑（已合并填报信息,含填报状态字段status）
    roi: gpd.GeoDataFrame，分区统计范围，有NAME字段区分不同分区
    '''
    gdf['area'] = gdf['area'] / 666.666
    df = pd.DataFrame([],columns=['范围','已填报养殖','已填报非养殖','未填报'])
    
    # 总体
    df.loc[0,'范围'] = '总体'
    for c in ['已填报养殖','已填报非养殖','未填报']:
        gdf1 = gdf.loc[gdf['status']==c, :]
        df.loc[0,c] = f'{len(gdf1)}个, {np.round(gdf1["area"].sum(),2)}亩'

    # 分区
    if roi is not None:
        for i,row in roi.iterrows():
            roii = row.geometry
            gdfi = gdf.loc[gdf.geometry.intersects(roii), :]
            
            df.loc[i+1,'范围'] = row['NAME']            
            for c in ['已填报养殖','已填报非养殖','未填报']:
                gdf1 = gdfi.loc[gdfi['status']==c, :]
                df.loc[i+1,c] = f'{len(gdf1)}个, {np.round(gdf1["area"].sum(),2)}亩'
    
    return df

def TBJDTJ01(df,outfile):
    '''
    全省填报进度统计-填报点
    '''
    cols = ['省','市','区县','街道','村委']
    dz = df['地址'].str.split('-',expand=True)
    for i in range(len(cols)):
        df[cols[i]] = dz.loc[:,i]

    with pd.ExcelWriter(outfile,engine='openpyxl') as writer:
        for i,c in enumerate(cols):
            tj = df.groupby(cols[0:i+1])[['养殖经营人名称']].count()
            tj.rename(columns={'养殖经营人名称':'总数'},inplace=True)

            t = df[df['填报状态']=='已填报养殖'].groupby(cols[0:i+1])[['养殖经营人名称']].count()
            tj.loc[t.index,'已填报养殖'] = t.values

            t = df[df['填报状态']=='已填报非养殖'].groupby(cols[0:i+1])[['养殖经营人名称']].count()
            tj.loc[t.index,'已填报非养殖'] = t.values
            tj['已填报'] = tj['已填报养殖'].fillna(0) + tj['已填报非养殖'].fillna(0)

            t = df[df['状态'].str.contains('校对')].groupby(cols[0:i+1])[['养殖经营人名称']].count()
            tj.loc[t.index,'在校对'] = t.values

            t = df[df['状态'].str.contains('通过')].groupby(cols[0:i+1])[['养殖经营人名称']].count()
            tj.loc[t.index,'已通过'] = t.values

            t = df[df['状态'].str.contains('返回')].groupby(cols[0:i+1])[['养殖经营人名称']].count()
            tj.loc[t.index,'已返回'] = t.values
            tj = tj.fillna(0)
            tj.to_excel(writer, sheet_name=c)
    
def TBJDTJ02(polygons,xzq,outfile,deletes=None):
    '''
    填报进度统计-池塘图斑总数，已填报的池塘图斑总数（含养殖和未使用状态）
    '''
    with pd.ExcelWriter(outfile,engine='openpyxl') as writer:
        # 市为单位
        tj2 = {
            '市':[],
            '总图斑个数':[],
            '已填报养殖个数':[],
            '已填报非养殖个数':[],
            '未填报个数':[],
            '总图斑面积(亩)':[],
            '已填报养殖面积(亩)':[],
            '已填报非养殖面积(亩)':[],
            '未填报面积(亩)':[],
            '备注':[]
        }
        shi = xzq['市'].unique()
        # if 'index_right' in polygons:
        #     polygons.drop(columns='index_right',inplace=True)
        for s in shi:
            # xzq_s = xzq[xzq['市']==s]
            # p0 = gpd.sjoin(polygons,xzq_s.loc[:,['geometry']]).drop_duplicates(subset=['geometry'])
            p0 = polygons[polygons['市']==s]
            p = p0[p0['Ndel']]
            tj2['市'].append(s)
            tj2['总图斑个数'].append(len(p))
            tj2['总图斑面积(亩)'].append(np.round(p["area"].sum()/666.666,2))
            p1 = p[p["填报状态"]=="已填报养殖"]
            tj2['已填报养殖个数'].append(len(p1))
            tj2['已填报养殖面积(亩)'].append(np.round(p1["area"].sum()/666.666,2))
            p2 = p[p["填报状态"]=="已填报非养殖"]
            tj2['已填报非养殖个数'].append(len(p2))
            tj2['已填报非养殖面积(亩)'].append(np.round(p2["area"].sum()/666.666,2))
            p3 = p[p["填报状态"]=="未填报"]
            tj2['未填报个数'].append(len(p3))
            tj2['未填报面积(亩)'].append(np.round(p3["area"].sum()/666.666,2))
            p4 = p0[p0['Ndel']==False]
            if len(p4)>0:
                tj2['备注'].append(f"{s}上报{len(p4)}个图斑删除，该部分不参与进度统计，后续抽样质控后删除")
            else:
                tj2['备注'].append("")
        tj2 = pd.DataFrame(tj2)
        i = len(tj2)
        tj2.loc[i,'市'] = '合计'
        tj2.loc[i,'总图斑个数':'未填报面积(亩)'] = tj2.loc[0:i,'总图斑个数':'未填报面积(亩)'].sum(axis=0)
        tj2.insert(5,'填报进度',((tj2['已填报养殖个数'] + tj2['已填报非养殖个数']) / tj2['总图斑个数']).apply(lambda x: format(x,".2%")))
        tj2.to_excel(writer, sheet_name='所有市', index=False)
        # writer.save()

        # 区为单位
        tj1 = {
            '市':[],
            '区县':[],
            '总图斑个数':[],
            '已填报养殖个数':[],
            '已填报非养殖个数':[],
            '未填报个数':[],
            '总图斑面积(亩)':[],
            '已填报养殖面积(亩)':[],
            '已填报非养殖面积(亩)':[],
            '未填报面积(亩)':[],
            '备注':[]
        }
        # if 'index_right' in polygons:
        #     polygons.drop(columns='index_right',inplace=True)
        for i,row in xzq.iterrows():
            # p0 = polygons[polygons.intersects(row.geometry)]
            p0 = polygons[(polygons['NAME']==row['NAME'])&(polygons['市']==row['市'])]
            p = p0[p0['Ndel']]
            tj1['市'].append(row['市'])
            tj1['区县'].append(row['NAME'])
            tj1['总图斑个数'].append(len(p))
            tj1['总图斑面积(亩)'].append(np.round(p["area"].sum()/666.666,2))
            p1 = p[p["填报状态"]=="已填报养殖"]
            tj1['已填报养殖个数'].append(len(p1))
            tj1['已填报养殖面积(亩)'].append(np.round(p1["area"].sum()/666.666,2))
            p2 = p[p["填报状态"]=="已填报非养殖"]
            tj1['已填报非养殖个数'].append(len(p2))
            tj1['已填报非养殖面积(亩)'].append(np.round(p2["area"].sum()/666.666,2))
            p3 = p[p["填报状态"]=="未填报"]
            tj1['未填报个数'].append(len(p3))
            tj1['未填报面积(亩)'].append(np.round(p3["area"].sum()/666.666,2))
            p4 = p0[p0['Ndel']==False]
            if len(p4)>0:
                tj1['备注'].append(f"{row['NAME']}上报{len(p4)}个图斑删除，该部分不参与进度统计，后续抽样质控后删除")
            else:
                tj1['备注'].append("")
        tj1 = pd.DataFrame(tj1)
        for s in xzq['市'].unique():
            t = tj1[tj1['市']==s].reset_index(drop=True)
            i = len(t)
            t.loc[i,'市'] = s
            t.loc[i,'区县'] = '合计'
            t.loc[i,'总图斑个数':'未填报面积(亩)'] = t.loc[0:i,'总图斑个数':'未填报面积(亩)'].sum(axis=0)
            t.insert(6,'填报进度',((t['已填报养殖个数'] + t['已填报非养殖个数']) / t['总图斑个数']).apply(lambda x: format(x,".2%")))
            t.to_excel(writer, sheet_name=s, index=False)
            # writer.save()

        # 上报删除数量统计
        

def TBJDTJ03(polygons,xzq,outfile):
    '''
    全省填报进度统计-已填报池塘图斑数，村级、乡镇级、县市区级分别完成校对上报的池塘图斑数量
    '''
    with pd.ExcelWriter(outfile,engine='openpyxl') as writer:        
        # 市为单位
        tj2 = {
            '市':[],
            '总图斑个数':[],
            '已填报图斑个数':[],
            '村已校对个数':[],
            '镇已校对个数':[],
            '县已校对个数':[],
            '总图斑面积(亩)':[],
            '已填报图斑面积(亩)':[],
            '村已校对面积(亩)':[],
            '镇已校对面积(亩)':[],
            '县已校对面积(亩)':[],
            '备注':[]
        }
        shi = xzq['市'].unique()
        # if 'index_right' in polygons:
        #     polygons.drop(columns='index_right',inplace=True)
        for s in shi:
            # xzq_s = xzq[xzq['市']==s]
            tj2['市'].append(s)

            # p0 = gpd.sjoin(polygons,xzq_s.loc[:,['geometry']]).drop_duplicates(subset=['geometry'])
            p0 = polygons[polygons['市']==s]
            p = p0[p0['Ndel']]
            tj2['总图斑个数'].append(len(p))
            tj2['总图斑面积(亩)'].append(np.round(p["area"].sum()/666.666,2))

            p = p[p["填报状态"]!="未填报"]  
            tj2['已填报图斑个数'].append(len(p))
            tj2['已填报图斑面积(亩)'].append(np.round(p["area"].sum()/666.666,2))

            p1 = p[p["状态值"]>1]
            tj2['村已校对个数'].append(len(p1))
            tj2['村已校对面积(亩)'].append(np.round(p1["area"].sum()/666.666,2))
            p2 = p[p["状态值"]>2]
            tj2['镇已校对个数'].append(len(p2))
            tj2['镇已校对面积(亩)'].append(np.round(p2["area"].sum()/666.666,2))
            p3 = p[p["状态值"]>3]
            tj2['县已校对个数'].append(len(p3))
            tj2['县已校对面积(亩)'].append(np.round(p3["area"].sum()/666.666,2))

            p4 = p0[p0['Ndel']==False]
            if len(p4)>0:
                tj2['备注'].append(f"{s}上报{len(p4)}个图斑删除，该部分不参与进度统计，后续抽样质控后删除")
            else:
                tj2['备注'].append("")

        tj2 = pd.DataFrame(tj2)
        i = len(tj2)
        tj2.loc[i,'市'] = '合计'
        tj2.loc[i,'总图斑个数':'县已校对面积(亩)'] = tj2.loc[0:i,'总图斑个数':'县已校对面积(亩)'].sum(axis=0)
        tj2.insert(6,'填报进度',(tj2['已填报图斑个数'] / tj2['总图斑个数']).apply(lambda x: format(x,".2%")))
        tj2.to_excel(writer, sheet_name='所有市', index=False)
        # writer.save()

        # 区为单位
        tj1 = {
            '市':[],
            '区县':[],
            '总图斑个数':[],
            '已填报图斑个数':[],
            '村已校对个数':[],
            '镇已校对个数':[],
            '县已校对个数':[],
            '总图斑面积(亩)':[],
            '已填报图斑面积(亩)':[],
            '村已校对面积(亩)':[],
            '镇已校对面积(亩)':[],
            '县已校对面积(亩)':[],
            '备注':[]
        }
        # if 'index_right' in polygons:
        #     polygons.drop(columns='index_right',inplace=True)
        for i,row in xzq.iterrows():
            tj1['市'].append(row['市'])
            tj1['区县'].append(row['NAME'])

            # p0 = polygons[polygons.intersects(row.geometry)]
            p0 = polygons[(polygons['NAME']==row['NAME'])&(polygons['市']==row['市'])].drop_duplicates(subset=['TBID'])
            p = p0[p0['Ndel']]
            tj1['总图斑个数'].append(len(p))
            tj1['总图斑面积(亩)'].append(np.round(p["area"].sum()/666.666,2))

            p = p[p["填报状态"]!="未填报"]  
            tj1['已填报图斑个数'].append(len(p))
            tj1['已填报图斑面积(亩)'].append(np.round(p["area"].sum()/666.666,2))
            p1 = p[p["状态值"]>1]
            tj1['村已校对个数'].append(len(p1))
            tj1['村已校对面积(亩)'].append(np.round(p1["area"].sum()/666.666,2))
            p2 = p[p["状态值"]>2]
            tj1['镇已校对个数'].append(len(p2))
            tj1['镇已校对面积(亩)'].append(np.round(p2["area"].sum()/666.666,2))
            p3 = p[p["状态值"]>3]
            tj1['县已校对个数'].append(len(p3))
            tj1['县已校对面积(亩)'].append(np.round(p3["area"].sum()/666.666,2))

            p4 = p0[p0['Ndel']==False]
            if len(p4)>0:
                tj1['备注'].append(f"{row['NAME']}上报{len(p4)}个图斑删除，该部分不参与进度统计，后续抽样质控后删除")
            else:
                tj1['备注'].append("")
        tj1 = pd.DataFrame(tj1)
        for s in xzq['市'].unique():
            t = tj1[tj1['市']==s].reset_index(drop=True)
            i = len(t)
            t.loc[i,'市'] = s
            t.loc[i,'区县'] = '合计'
            t.loc[i,'总图斑个数':'县已校对面积(亩)'] = t.loc[0:i,'总图斑个数':'县已校对面积(亩)'].sum(axis=0)
            t.insert(7,'填报进度',(t['已填报图斑个数'] / t['总图斑个数']).apply(lambda x: format(x,".2%")))
            t.to_excel(writer, sheet_name=s, index=False)
            # writer.save()

def TBJDTJ04(polygons,fields,outfile,subtotle=True):
    '''
    按字段统计数量/进度
    '''
    ndel = polygons['Ndel']
    ytb = polygons['填报状态']!='未填报'
    cjd = polygons['状态值']>1
    zjd = polygons['状态值']>2
    xjd = polygons['状态值']>3
    sc = polygons['Ndel']==False

    polygons.loc[ndel,'总图斑'] = polygons.loc[ndel,'area']*0.0015
    polygons.loc[ndel & ytb,'已填报'] = polygons.loc[ndel & ytb,'area']*0.0015
    polygons.loc[ndel & cjd,'村已校对'] = polygons.loc[ndel & cjd,'area']*0.0015
    polygons.loc[ndel & zjd,'镇已校对'] = polygons.loc[ndel & zjd,'area']*0.0015
    polygons.loc[ndel & xjd,'县已校对'] = polygons.loc[ndel & xjd,'area']*0.0015
    polygons.loc[sc,'备注'] = polygons.loc[sc,'area']*0.0015

    grouped = polygons.groupby(fields)
    # 数量统计
    t1 = grouped[['总图斑','已填报','村已校对','镇已校对','县已校对']].count().rename(columns={
        '总图斑':'总图斑个数',
        '已填报':'已填报个数',
        '村已校对':'村已校对个数',
        '镇已校对':'镇已校对个数',
        '县已校对':'县已校对个数'
    })
    # 面积统计
    t2 = grouped[['总图斑','已填报','村已校对','镇已校对','县已校对']].sum().rename(columns={
        '总图斑':'总图斑面积(亩)',
        '已填报':'已填报面积(亩)',
        '村已校对':'村已校对面积(亩)',
        '镇已校对':'镇已校对面积(亩)',
        '县已校对':'县已校对面积(亩)'
    })
    # 上报删除统计
    t3 = grouped[['备注']].count()

    # 区、市统计表
    tj_qx = pd.concat([t1,t2,t3],axis=1)
    
    # 小计
    if subtotle:
        cols = tj_qx.columns
        tj_shi = tj_qx.groupby(level=list(range(len(fields)-1)))[cols].sum()
        for i in tj_shi.index:
            if isinstance(i,tuple):
                idx = list(i)
            else:
                idx=[i]
            idx.append('a')
            idx = tuple(idx)
            tj_qx.loc[idx,cols] = tj_shi.loc[i,cols]
        tj_qx = tj_qx.sort_index(ascending=False)
        # 替换'a'为小计
        idx_name = list(tj_qx.index.names)
        for i,c in enumerate(idx_name):
            if i < (len(idx_name)-1):
                tj_qx[c] = tj_qx.index.get_level_values(i)
            else:
                tj_qx[c] = tj_qx.index.get_level_values(i).str.replace('a','合计')
        tj_qx = tj_qx.set_index(idx_name)
        # 合计的总表
        idx_name = list(tj_shi.index.names)
        if len(idx_name)==1:
            tj_shi.loc['合计',cols] = tj_shi.sum(axis=0)
        else:
            idx = []
            for i in range(len(idx_name)-1):
                idx.append(tj_shi.index.get_level_values(i).unique()[0]) 
            idx.append('合计')
            tj_shi.loc[tuple(idx),cols] = tj_shi.sum(axis=0)
        tj_shi.insert(5,'填报进度',(tj_shi['已填报个数'] / tj_shi['总图斑个数']).apply(lambda x: format(x,".2%")))
        idx = tj_shi['备注'] != 0
        tj_shi.loc[idx,'备注'] = '上报' + tj_shi.loc[idx,'备注'].astype('int').astype('str') + '个图斑删除，该部分不参与进度统计，后续抽样质控后删除'
        tj_shi.loc[~idx,'备注'] = ''
        tj_shi.to_excel(outfile.replace('.xlsx','-合计.xlsx'))

    # 填报进度
    tj_qx.insert(5,'填报进度',(tj_qx['已填报个数'] / tj_qx['总图斑个数']).apply(lambda x: format(x,".2%")))
    
    # 备注删除数量
    idx = tj_qx['备注'] != 0
    tj_qx.loc[idx,'备注'] = '上报' + tj_qx.loc[idx,'备注'].astype('int').astype('str') + '个图斑删除，该部分不参与进度统计，后续抽样质控后删除'
    tj_qx.loc[~idx,'备注'] = ''

    # 导出统计表
    tj_qx.to_excel(outfile)

    # 导出未填报
    wtb = polygons[polygons['填报状态']=='未填报']
    wtb = wtb[wtb['Ndel']]
    wtb['TBID'] = wtb['TBID'].str.replace(',','')
    wtb['地址'] = wtb[fields[0]]
    for i in range(1,len(fields)):
        wtb['地址'] += wtb[fields[i]]
    wtb = wtb.sort_values(by='地址')
    wtb.set_index(fields,inplace=True)
    wtb.loc[:,['TBID']].to_excel(outfile.replace('.xlsx','-未填报图斑.xlsx'))

def TBJDTJ05(polygons,fields,outfile,subtotle=True):
    '''
    按字段统计数量/进度
    '''
    ndel = polygons['Ndel']
    tbyz = polygons['填报状态']=='已填报养殖'
    tbfyz = polygons['填报状态']=='已填报非养殖'
    wtb = polygons['填报状态']=='未填报'
    sc = polygons['Ndel']==False

    polygons.loc[ndel,'总图斑'] = polygons.loc[ndel,'area']*0.0015
    polygons.loc[ndel & tbyz,'已填报养殖'] = polygons.loc[ndel & tbyz,'area']*0.0015
    polygons.loc[ndel & tbfyz,'已填报非养殖'] = polygons.loc[ndel & tbfyz,'area']*0.0015
    polygons.loc[ndel & wtb,'未填报'] = polygons.loc[ndel & wtb,'area']*0.0015
    polygons.loc[sc,'备注'] = polygons.loc[sc,'area']*0.0015

    grouped = polygons.groupby(fields)
    # 数量统计
    t1 = grouped[['总图斑','已填报养殖','已填报非养殖','未填报']].count().rename(columns={
        '总图斑':'总图斑个数',
        '已填报养殖':'已填报养殖个数',
        '已填报非养殖':'已填报非养殖个数',
        '未填报':'未填报个数',
    })
    # 面积统计
    t2 = grouped[['总图斑','已填报养殖','已填报非养殖','未填报']].sum().rename(columns={
        '总图斑':'总图斑面积(亩)',
        '已填报养殖':'已填报养殖面积(亩)',
        '已填报非养殖':'已填报非养殖面积(亩)',
        '未填报':'未填报面积(亩)',
    })
    # 上报删除统计
    t3 = grouped[['备注']].count()

    # 区、市统计表
    tj_qx = pd.concat([t1,t2,t3],axis=1)
    
    # 小计
    if subtotle:
        cols = tj_qx.columns
        tj_shi = tj_qx.groupby(level=list(range(len(fields)-1)))[cols].sum()
        for i in tj_shi.index:
            if isinstance(i,tuple):
                idx = list(i)
            else:
                idx=[i]
            idx.append('a')
            idx = tuple(idx)
            tj_qx.loc[idx,cols] = tj_shi.loc[i,cols]
        tj_qx = tj_qx.sort_index(ascending=False)
        # 替换'a'为小计
        idx_name = list(tj_qx.index.names)
        for i,c in enumerate(idx_name):
            if i < (len(idx_name)-1):
                tj_qx[c] = tj_qx.index.get_level_values(i)
            else:
                tj_qx[c] = tj_qx.index.get_level_values(i).str.replace('a','合计')
        tj_qx = tj_qx.set_index(idx_name)
        # 合计的总表
        idx_name = list(tj_shi.index.names)
        if len(idx_name)==1:
            tj_shi.loc['合计',cols] = tj_shi.sum(axis=0)
        else:
            idx = []
            for i in range(len(idx_name)-1):
                idx.append(tj_shi.index.get_level_values(i).unique()[0]) 
            idx.append('合计')
            tj_shi.loc[tuple(idx),cols] = tj_shi.sum(axis=0)
        tj_shi.insert(4,'填报进度',((tj_shi['已填报养殖个数']+tj_shi['已填报非养殖个数']) / tj_shi['总图斑个数']).apply(lambda x: format(x,".2%")))
        idx = tj_shi['备注'] != 0
        tj_shi.loc[idx,'备注'] = '上报' + tj_shi.loc[idx,'备注'].astype('int').astype('str') + '个图斑删除，该部分不参与进度统计，后续抽样质控后删除'
        tj_shi.loc[~idx,'备注'] = ''
        tj_shi.to_excel(outfile.replace('.xlsx','-合计.xlsx'))

    # 填报进度
    tj_qx.insert(4,'填报进度',((tj_qx['已填报养殖个数']+tj_qx['已填报非养殖个数']) / tj_qx['总图斑个数']).apply(lambda x: format(x,".2%")))
    
    # 备注删除数量
    idx = tj_qx['备注'] != 0
    tj_qx.loc[idx,'备注'] = '上报' + tj_qx.loc[idx,'备注'].astype('int').astype('str') + '个图斑删除，该部分不参与进度统计，后续抽样质控后删除'
    tj_qx.loc[~idx,'备注'] = ''

    # 导出统计表
    tj_qx.to_excel(outfile)


def zongbiao(xlsxfile):
    dfs = pd.read_excel(xlsxfile,sheet_name=None)
    del dfs['所有市']
    df = pd.concat(dfs.values(),ignore_index=True)
    df = df.set_index(['市','区县'])
    df.to_excel(xlsxfile.replace('.xlsx','-各区县汇总.xlsx'))

def exportTablesByFields(df,groupby,outpth,fields=None):
    '''
    按字段分表文件导出
    '''
    # 统计组别
    df = df.set_index(groupby)
    groups = df.index.unique()
    
    # 提取字段
    if fields is not None:
        df = df[fields]

    # 导出数据表
    for g in groups:
        df.loc[g,:].to_excel(f"{outpth}\\{'-'.join(g)}.xlsx")

def exportTablesByFieldsAndDFXZQ(gdf,groupby,outpth,dfxzqfile,fields=None):
    '''
    按字段分表文件导出
    '''
    if 'index_right' in gdf.columns:
        gdf = gdf.drop(['index_right'], axis=1)
    gdf = gdf.set_index('TBID')

    # 地方行政区划
    if dfxzqfile.endswith('.gpkg'):
        dfxzq = gpd.read_file(dfxzqfile)
    else:
        dfxzq = dfxzqfile
    dfxzq = dfxzq.to_crs(gdf.crs)
    dfxzq['地方区划'] = dfxzq['地方市'] + '-' + dfxzq['地方区县'] + '-' + dfxzq['镇名称']
    # 按镇判断相交，添加地方行政区信息
    print(f"按镇判断相交，添加地方行政区信息")
    gdf_sjoin = gpd.sjoin(gdf,dfxzq)
    print("sjoin finished")
    gdf_sjoin_tbid_group = gdf_sjoin.groupby('TBID')['地方区划'].unique().map(lambda x: ';'.join(x))
    print("sjoin_tbid_group finished")
    idx = gdf_sjoin_tbid_group.index
    gdf.loc[idx,'地方区划'] = gdf_sjoin_tbid_group[idx]
    gdf = gdf.sort_index()
    print("地方区划 finished")
    
    # 统计组别
    gdf['TBID'] = gdf.index
    gdf['TBID'] = gdf['TBID'].str.replace(',','')
    gdf = gdf.set_index(groupby)

    # 提取字段
    if fields is not None:
        fields.append('地方区划')
        gdf = gdf[fields]    

    # 导出数据表
    if outpth.endswith('xlsx'):
        gdf.to_excel(outpth)
    else:
        os.makedirs(outpth,exist_ok=True)
        groups = gdf.index.unique()
        for g in groups:
            print(f"导出：{'-'.join(g)}")

############################################################################
############################## 网页地图 #####################################
############################################################################
def createYDMap(yd1,yd2,yd3):
    '''
    生成疑点分布图
    yd1: points 用户填报点
    yd2: polygons 未填报图斑
    yd3: polygons 重复填报图斑
    xzq: polygons 行政区划
    '''
    minx,miny,maxx,maxy = yd1.total_bounds
    centroid_coords = [(miny+maxy)/2,(minx+maxx)/2]
    m = folium.Map(location=centroid_coords, zoom_start=15,tiles=None)
    add_tian_di_tu_layers(m)

    pt_cluster = folium.plugins.MarkerCluster().add_to(m)
    txt_cluster = folium.plugins.MarkerCluster().add_to(m)

    folium.GeoJson(
        yd1,
        name='填报疑点',
    ).add_to(pt_cluster)

    for i, row in yd1.iterrows():
        folium.Marker(
            location=[row['longitude'],row['latitude']],
            show='True',
            icon=folium.DivIcon(
                icon_size=(30, 10),
                icon_anchor=(0, 0),
                html=f'''
                    <div style="font-size: 12pt; color: black; text-shadow:
                        -1px -1px 0 #fff,
                        1px -1px 0 #fff,
                        -1px 1px 0 #fff,
                        1px 1px 0 #fff;">
                        {i}
                    </div>'''
            )
        ).add_to(txt_cluster)

    style_function = lambda x: {
        'fillColor': '#transparent', 
        'color': '#00FF00',  # 绿色
        'weight': 1,
    }
    folium.GeoJson(
        yd2,
        style_function=style_function,
        name='未填报图斑',
    ).add_to(m)

    style_function = lambda x: {
        'fillColor': '#transparent',
        'color': '#FF0000',  # 红色
        'weight': 1,
    }
    folium.GeoJson(
        yd3,
        style_function=style_function,
        name='填报多点的图斑',
    ).add_to(m)

    # folium.LayerControl().add_to(m)

    return m

def createCTMap(ct,xzq,labelfield='TBID'):
    '''
    创建池塘地图
    '''
    try:
        minx,miny,maxx,maxy = xzq.total_bounds
    except:
        minx,miny,maxx,maxy = xzq.bounds
    centroid_coords = [(miny+maxy)/2,(minx+maxx)/2]
    m = folium.Map(location=centroid_coords, zoom_start=15,tiles=None)
    add_tian_di_tu_layers(m)

    # txt_cluster = folium.plugins.MarkerCluster().add_to(m)

    for i, row in ct.iterrows():
        folium.Marker(
            location=[row['latitude'],row['longitude']],
            show='True',
            icon=folium.DivIcon(
                icon_size=(30, 10),
                icon_anchor=(0, 0),
                html=f'''
                    <div style="font-size: 12pt; color: black; text-align:center;text-shadow:
                        -1px -1px 0 #fff,
                        1px -1px 0 #fff,
                        -1px 1px 0 #fff,
                        1px 1px 0 #fff;">
                        {row[labelfield].replace(',','')}
                    </div>'''
            )
        ).add_to(m)

    style_function = lambda x: {
        'fillColor': '#transparent', 
        'color': '#FFFF00',  # 黄色
        'weight': 1,
    }
    folium.GeoJson(
        ct,
        style_function=style_function,
        name='池塘图斑',
    ).add_to(m)

    style_function = lambda x: {
        'fillColor': '#transparent',
        # 'color': '#000000',  # 黑色
        'color': '#FF0000',  # 红色
        'weight': 2,
    }
    folium.GeoJson(
        xzq,
        style_function=style_function,
        name='行政区',
    ).add_to(m)

    folium.LayerControl().add_to(m)

    return m

def createCTMap2(ct,xzq):
    '''
    创建池塘地图-加聚合
    '''
    try:
        minx,miny,maxx,maxy = xzq.total_bounds
    except:
        minx,miny,maxx,maxy = xzq.bounds
    centroid_coords = [(miny+maxy)/2,(minx+maxx)/2]
    m = folium.Map(location=centroid_coords, zoom_start=15,tiles=None)
    add_tian_di_tu_layers(m)

    txt_cluster = folium.plugins.MarkerCluster().add_to(m)

    for i, row in ct.iterrows():
        folium.Marker(
            location=[row['latitude'],row['longitude']],
            show='True',
            icon=folium.DivIcon(
                icon_size=(30, 10),
                icon_anchor=(0, 0),
                html=f'''
                    <div style="font-size: 12pt; color: black; text-align:center;text-shadow:
                        -1px -1px 0 #fff,
                        1px -1px 0 #fff,
                        -1px 1px 0 #fff,
                        1px 1px 0 #fff;">
                        {row['TBID'].replace(',','')}
                    </div>'''
            )
        ).add_to(txt_cluster)

    style_function = lambda x: {
        'fillColor': '#transparent', 
        'color': '#FFFF00',  # 黄色
        'weight': 1,
    }
    folium.GeoJson(
        ct,
        style_function=style_function,
        name='池塘图斑',
    ).add_to(m)

    style_function = lambda x: {
        'fillColor': '#transparent',
        'color': '#000000',  # 黑色
        'weight': 2,
    }
    folium.GeoJson(
        xzq,
        style_function=style_function,
        name='行政区',
    ).add_to(m)

    folium.LayerControl().add_to(m)

    return m

def add_tian_di_tu_layers(map_object):
    tian_di_tu_normal_map = ("https://t6.tianditu.gov.cn/img_w/wmts"
                         "?SERVICE=WMTS&REQUEST=GetTile&VERSION=1.0.0"
                         "&LAYER=img&STYLE=default&TILEMATRIXSET=w"
                         "&FORMAT=tiles&TILECOL={x}&TILEROW={y}"
                         "&TILEMATRIX={z}&tk=5625113a2addc9a7594d0fffe3811311")

    # 天地图注记图层的URL模板
    tian_di_tu_zhuji = ("https://t6.tianditu.gov.cn/cia_w/wmts"
                        "?SERVICE=WMTS&REQUEST=GetTile&VERSION=1.0.0"
                        "&LAYER=cia&STYLE=default&TILEMATRIXSET=w"
                        "&FORMAT=tiles&TileCol={x}&TileRow={y}"
                        "&TileMatrix={z}&tk=5625113a2addc9a7594d0fffe3811311")
    
    for name, url in [("天地图影像", tian_di_tu_normal_map), ("天地图注记", tian_di_tu_zhuji)]:
        folium.TileLayer(
            tiles=url,
            attr=name,
            name=name,
            overlay=name,
            control=False
        ).add_to(map_object)


def generateMapByCounties(ct,xzq,cun_col,zhen_col,dm_col,xlsfile,outpath):
    '''
    ct: gpd.GeoDataFrame 池塘图斑
    xzq: gpd.GeoDataFrame 行政区图斑
    cun_col: str 行政区中记录村名称的列名
    zhen_col: str 行政区中记录镇名称的列名
    '''
    ct.drop_duplicates(subset=['TBID'],inplace=True)
    ct['longitude'] = ct.geometry.centroid.x
    ct['latitude'] = ct.geometry.centroid.y
    df = pd.DataFrame()
    for i,row in xzq.iterrows():
        geom = row.geometry
        intersects = ct[ct.intersects(geom)]
        df.loc[i,'镇'] = row[zhen_col]
        df.loc[i,'村'] = row[cun_col]
        df.loc[i,'行政村代码'] = row[dm_col]
        df.loc[i,'池塘数量'] = len(intersects)
        intersects_map = createCTMap(intersects,geom)
        os.makedirs(row[zhen_col],exist_ok=True)
        intersects_map.save(f'{outpath}\\{row[zhen_col]}\\{row[cun_col]}-{row[dm_col]}-池塘分布.html')

        workbook = openpyxl.load_workbook(xlsfile)
        sheet = workbook.active
        for j,v in enumerate(intersects['TBID'].values):
            sheet.cell(6+j,3).value = v.replace(',','')
        workbook.save(f'{outpath}\\{row[zhen_col]}\\{row[cun_col]}-{row[dm_col]}-池塘信息表.xlsx')
    df.to_excel(f'{outpath}\\池塘数量统计.xlsx')

def generateMapByDistrict(ct,xzq,xlsfile,outpath):
    '''
    按行政区生成地图网页和表格
    ct: gpd.GeoDataFrame 池塘图斑
    xzq: gpd.GeoDataFrame 行政区图斑
    '''
    ct.drop_duplicates(subset=['TBID'],inplace=True)
    ct['longitude'] = ct.geometry.centroid.x
    ct['latitude'] = ct.geometry.centroid.y
    for i,row in xzq.iterrows():
        geom = row.geometry
        intersects = ct[ct.intersects(geom)]
        intersects_map = createCTMap(intersects,geom)
        intersects_map.save(f'{outpath}\\池塘分布.html')

        workbook = openpyxl.load_workbook(xlsfile)
        sheet = workbook.active
        for j,v in enumerate(intersects['TBID'].values):
            sheet.cell(6+j,3).value = v.replace(',','')
        workbook.save(f'{outpath}\\池塘信息表.xlsx')

def generateMapByDistrict2(ct,xzq,xlsfile,outpath):
    '''
    按行政区生成地图网页和表格-加聚合
    ct: gpd.GeoDataFrame 池塘图斑
    xzq: gpd.GeoDataFrame 行政区图斑
    '''
    ct.drop_duplicates(subset=['TBID'],inplace=True)
    ct['longitude'] = ct.geometry.centroid.x
    ct['latitude'] = ct.geometry.centroid.y
    for i,row in xzq.iterrows():
        geom = row.geometry
        intersects = ct[ct.intersects(geom)]
        intersects_map = createCTMap2(intersects,geom)
        intersects_map.save(f'{outpath}\\池塘分布-聚合.html')

        # workbook = openpyxl.load_workbook(xlsfile)
        # sheet = workbook.active
        # for j,v in enumerate(intersects['TBID'].values):
        #     sheet.cell(6+j,3).value = v.replace(',','')
        # workbook.save(f'{outpath}\\池塘信息表.xlsx')


def isDulplicate(gdf,cols):
    '''
    检查矢量有无重复
    gdf: gpd.GeoDataFrame 
    cols: 检查列 str or [str]
    '''
    gdf1 = gdf.dissolve(by=cols)

    if len(gdf1)==len(gdf):
        return False
    else:
        return True
    
def XQZTable(gdf,dm_col,cun_col,zhen_col):
    '''
    写出行政区划名称表
    gdf: gpd.GeoDataFrame 
    dm_col ： str 记录村行政代码的列名
    cun_col ： str 记录村名称的列名
    zhen_col ： str 记录镇名称的列名
    '''
    df = gdf.loc[:,[zhen_col,cun_col,dm_col]]
    df = df.rename(columns={
        zhen_col:'镇',
        cun_col:'村',
        dm_col:'行政区代码'
    })

    return df


############################################################################
############################### 其他 ######################################
############################################################################
def readZDSM(ZDSM_FILE):
    '''
    软件数据库字段 英-中对照 字典
    '''
    with open(ZDSM_FILE,encoding='utf-8') as f:
        data = f.read().split('\n')[1:-2]
        eng = [d.split("\'")[0].replace('`','').strip() for d in data]
        chn = [d.split("\'")[1].split(' ')[0] for d in data]
    zd_dict = {}
    for i in range(len(chn)):
        zd_dict[eng[i]] = chn[i]

    return zd_dict

def toExcelByQZX(outpath,df,field):
    '''
    按字段唯一值分sheet输出表格
    '''
    # 总表
    
    df.to_excel(os.path.join(outpath,'总表.xlsx'),sheet_name='总表')

    xzq = np.unique(df[field].values)
    xzq = ['-'.join(x.split('-')[0:-1]) for x in xzq]
    xzq = list(set(xzq))
    address = df[field].str.split('-',expand=True)
    address = address[0] + '-' + address[1] + '-' + address[2] + '-' + address[3]
    for x in xzq:
        subsets = df[address==x]
        subsets.to_excel(os.path.join(outpath,f'{x}.xlsx'))

def toExcelByQZX2(outpath,df,field):
    '''
    按字段唯一值分sheet输出表格,某疑点列全为无异常则该列删除
    '''
    yd_columns = ['名称疑点','位置疑点','水面面积疑点','合同面积疑点','池塘合并疑点','亩产量疑点']
    df.to_excel(os.path.join(outpath,'总表.xlsx'),sheet_name='总表')
    xzq = np.unique(df[field].values)
    xzq = ['-'.join(x.split('-')[0:-1]) for x in xzq]
    xzq = list(set(xzq))
    address = df[field].str.split('-',expand=True)
    address = address[0] + '-' + address[1] + '-' + address[2] + '-' + address[3]
    for x in xzq:
        subsets = df[address==x]
        for c in yd_columns:
            v = np.unique(subsets[c].values)
            if (len(v) == 1) & (v[0] == '无异常'):
                subsets = subsets.drop(columns=[c])
        subsets.to_excel(os.path.join(outpath,f'{x}.xlsx'))

# 删除异常填报养殖名称的条目，合并同类养殖名称的条目
def clean_data_by_variety(df):
    # 剔除异常填报品种的条目
    df3 = df.copy()
    for var in df3.loc[:,'variety_name'].unique():
        if len(var.split(',')) > len(set(var.split(','))):
            # print(var)
            # print(df3.loc[df3['名称'] == var].index)
            df3.drop(index=df3.loc[df['variety_name'] == var].index, inplace=True)
    
    # 生成主要养殖品种（去重）
    var_seq = df3.loc[:,'variety_name'].unique()
    var_freq = np.zeros(len(var_seq))

    for i in range(len(var_seq) - 1):
        for j in range(i + 1, len(var_seq)):
            if set(var_seq[i].split(',')) == set(var_seq[j].split(',')):
                var_freq[j] = var_freq[j] + 1
    var_unique = var_seq[var_freq == 0]
    
    # 增加一列，存放修改后的养殖品种和亩产量
    df3['养殖品种修正'] = ''
    df3['亩产量修正'] = ''
    for i in df3.index:
        var0 = df3.loc[i, 'variety_name']
        var_vol0 = df3.loc[i, 'yield']
        for var in var_unique:
            if (set(var0.split(',')) == set(var.split(','))):
                v0_seq = np.array(var0.split(','))
                vq_seq = np.array(var.split(','))
                vol0_seq = np.array(var_vol0.split(','))
                var_vol1 = ''
                for vq in vq_seq:
                    var_vol1 = var_vol1 + str(vol0_seq[v0_seq == vq][0]) + ','
                var_vol1 = var_vol1[:-1]

                df3.loc[i, '养殖品种修正'] = var
                df3.loc[i, '亩产量修正'] = var_vol1

    df3['variety_name'] = df3['养殖品种修正']
    df3['yield'] = df3['亩产量修正']

    return df3.drop(columns=['养殖品种修正','亩产量修正'])


# 按地方上报、双方确认的内容调整图斑权属
def reallocatePolygons(gdf,lkt):
    '''
    gdf: 池塘图斑
    lkt: pd.DataFrame 包括列：TBID	市	区县	上报单位 
    '''
    gdf['new_tbid'] = gdf['TBID'].str.replace(',','')
    gdf.set_index('new_tbid',inplace=True) 
    idx = lkt['TBID'].values
    lkt.set_index('TBID',inplace=True)
    
    try:
        gdf.loc[idx,'市'] = lkt.loc[idx,'市']
        gdf.loc[idx,'区县'] = lkt.loc[idx,'区县']
    except:
        gdf.loc[idx,'地方市'] = lkt['市']
        gdf.loc[idx,'地方区县'] = lkt['区县']
    
    gdf.reset_index(inplace=True,drop=True)
    
    return gdf

def Shrimp_yieldAnayesis(df):
    '''
    单养虾蟹塘面积疑点：虾塘、蟹塘面积累积概率分布大于0.95作为疑点
    单养品种亩产量疑点：对填报次数达100及以上的单养品种作此疑点判定,每品种亩产量累积概率分布大于0.95作为疑点
    '''
    df['亩产量疑点'] = ''
    df['虾蟹塘面积疑点'] = ''
    df['养殖品种/预计亩产量'] = df['养殖品种/预计亩产量'].str.replace('斤/亩','')
    idx = df[df['养殖品种/预计亩产量']!='/'].index
    for i in idx:
        yzxx = df.loc[i,'养殖品种/预计亩产量'].split('，')
        pz = np.array([y.split(':')[0] for y in yzxx])
        cl = np.array([y.split(':')[1] for y in yzxx])
        sortedidx = np.argsort(pz)
        df.loc[i,'养殖品种'] = ','.join(pz[sortedidx])
        df.loc[i,'亩产量'] = ','.join(cl[sortedidx])
    
    DYPZ=df.loc[idx,:]
    DYPZ=DYPZ[DYPZ['养殖品种'].apply(lambda x: len(x.split(',')) == 1)]
    new=DYPZ.copy()
    DYPZ=DYPZ[DYPZ['图斑面积'].notnull()]
    idx_hexie=DYPZ['养殖品种']== '河蟹'
    idx_xia=DYPZ['养殖品种'].str.contains('虾')
    hexie = DYPZ.loc[idx_hexie,:]
    xia = DYPZ.loc[idx_xia,:]

    # 设定面积异常值
    y1 = np.arange(1, len(hexie['图斑面积']) + 1) / len(hexie['图斑面积'])
    y2 = np.arange(1, len(xia['图斑面积']) + 1) / len(xia['图斑面积'])
    xie=np.sort(hexie['图斑面积'])
    xia_area=np.sort(xia['图斑面积'])
    S1=0.95
    S2=0.05
    idx1=np.searchsorted(y1, S1)
    idx11=np.searchsorted(y2, S1)
    areamax1 = xie[idx1]
    areamax2 = xia_area[idx11]
    idx_H1 = hexie['图斑面积']>=areamax1
    idx_X1 = xia['图斑面积']>=areamax2
    hexie.loc[idx_H1,'虾蟹塘面积疑点'] = '蟹塘面积过大'
    xia.loc[idx_X1,'虾蟹塘面积疑点'] = '虾塘面积过大'
    new_df=pd.concat([xia['虾蟹塘面积疑点'],hexie['虾蟹塘面积疑点']])
    df.loc[idx,'虾蟹塘面积疑点'] = new_df

    YZPZ_MCL=new[new['亩产量']!='0.00']
    types=YZPZ_MCL['养殖品种'].unique()
    for i in types:
        if len(YZPZ_MCL[YZPZ_MCL['养殖品种']==i])>=100:
            DYPZ_MCL=YZPZ_MCL[YZPZ_MCL['养殖品种']==i]
            DYPZ_MCL['亩产量'] = [float(num) for num in DYPZ_MCL['亩产量']]
            data=np.sort(DYPZ_MCL['亩产量'])
            y = np.arange(1, len(data) + 1) / len(data)
            # 设定异常值
            # V1=0.05
            V2=0.95
            upidx = np.searchsorted(y, V2)
            upper_outliers = data[upidx]
            indices = DYPZ_MCL[DYPZ_MCL['亩产量'] >upper_outliers].index
            df.loc[indices,'亩产量疑点'] = '亩产量过大'

    return df

def extractYZPZ(df):
    '''
    提取养殖品种
    '''
    print("check df.index:")
    print(f"总长{len(df)},最大索引{df.index.max()}")
    yzxx = df['养殖品种/预计亩产量'].str.replace('斤/亩','')
    yzxx = yzxx.str.split('，',expand=True)
    yzxx = yzxx.fillna('/')
    yzpz = pd.DataFrame(columns=yzxx.columns,index=yzxx.index)
    mcl = pd.DataFrame(columns=yzxx.columns,index=yzxx.index)
    print(f"养殖品种、亩产量拆分")
    for c in yzxx.columns:
        idx = yzxx[c]!='/'
        yzpz.loc[idx,c] = yzxx.loc[idx,c].str.split(':',expand=True)[0]
        mcl.loc[idx,c] = yzxx.loc[idx,c].str.split(':',expand=True)[1]
    
    yzpz = yzpz.fillna('/')
    mcl = mcl.fillna(0)
    mcl = mcl.to_numpy().astype('float')
    yzpz_unq = np.unique(yzpz.to_numpy())
    n = yzpz_unq.shape[0]-1
    print(f"共{n}个品种")
    for i,pz in enumerate(yzpz_unq[yzpz_unq!='/']):
        print(f"{i+1}/{n}:{pz}")
        pz_idx = np.argwhere(yzpz==pz)
        df.loc[df.index[pz_idx[:,0]],f'{pz}亩产量(斤/亩)'] = mcl[yzpz==pz]

    return df