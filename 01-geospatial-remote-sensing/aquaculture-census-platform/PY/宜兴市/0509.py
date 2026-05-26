"""
项目名称: aquaculture-census-platform
技术领域: 01-geospatial-remote-sensing
模块说明: 0509.py - 核心业务算法实现
作者: 杨佳 (资深 AI 算法与遥感工程师)
"""

from __future__ import annotations

import math
import re
from pathlib import Path

import geopandas as gpd
import pandas as pd
from openpyxl import load_workbook


GROUP_COLS = ["所在镇村（镇街+村）", "养殖主体姓名（名称）"]
ID_COL = "图斑编号"
AREA_COL = "塘口面积（亩）"
TOTAL_AREA_COL = "养殖户总面积（亩）"
COUNT_COL = "养殖户塘口数量（当前镇村）"


def normalize_text(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "nat"}:
        return ""
    return text


def parse_area(value) -> float | None:
    if pd.isna(value):
        return None
    if isinstance(value, (int, float)) and not pd.isna(value):
        return float(value)
    text = normalize_text(value)
    if not text:
        return None
    cleaned = re.sub(r"[^\d.]+", "", text)
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def choose_group_total(values: list[float | None]) -> float:
    numeric = [float(v) for v in values if v is not None and not math.isnan(v)]
    if not numeric:
        return 0.0
    counts = pd.Series(numeric).round(6).value_counts()
    top_count = counts.iloc[0]
    candidates = sorted(counts[counts == top_count].index.tolist())
    return float(candidates[-1])


def detect_header_row(excel_path: Path) -> int:
    wb = load_workbook(excel_path, read_only=True, data_only=False)
    ws = wb[wb.sheetnames[0]]
    row1 = [ws.cell(row=1, column=i).value for i in range(1, min(ws.max_column, 40) + 1)]
    row2 = [ws.cell(row=2, column=i).value for i in range(1, min(ws.max_column, 40) + 1)]
    wb.close()

    row1_text = {normalize_text(v) for v in row1 if normalize_text(v)}
    row2_text = {normalize_text(v) for v in row2 if normalize_text(v)}
    if {"基本情况", "基础信息", "养殖情况", "治理情况"} & row1_text and {ID_COL, GROUP_COLS[1]} <= row2_text:
        return 1
    return 0


def load_excel_df(excel_path: Path) -> pd.DataFrame:
    df = pd.read_excel(excel_path, header=detect_header_row(excel_path))
    df.columns = [normalize_text(c) for c in df.columns]
    df = df.loc[:, [c for c in df.columns if c and not c.startswith("Unnamed:")]]
    return df


def build_updated_dataframe(excel_df: pd.DataFrame, pond_gdf: gpd.GeoDataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    df = excel_df.copy()
    df[ID_COL] = df[ID_COL].astype(str).str.strip()
    for col in GROUP_COLS:
        df[col] = df[col].apply(normalize_text)
    df["_group_total_area"] = df[TOTAL_AREA_COL].apply(parse_area)

    pond = pond_gdf.copy()
    pond["tbid"] = pond["tbid"].astype(str).str.strip()
    pond = pond.to_crs("EPSG:32650")
    pond["_plot_area_mu"] = pond.geometry.area * 0.0015

    area_df = pd.DataFrame(pond[["tbid", "_plot_area_mu"]])
    df = df.merge(area_df, left_on=ID_COL, right_on="tbid", how="left")
    df["_matched"] = df["tbid"].notna()

    issues: list[dict] = []
    group_total_map = df.groupby(GROUP_COLS)["_group_total_area"].apply(lambda s: choose_group_total(s.tolist())).to_dict()
    df[TOTAL_AREA_COL] = df.apply(lambda r: round(group_total_map[(r[GROUP_COLS[0]], r[GROUP_COLS[1]])], 2), axis=1)
    df[COUNT_COL] = df.groupby(GROUP_COLS)[ID_COL].transform("count")
    df["_group_plot_area_sum"] = df.groupby(GROUP_COLS)["_plot_area_mu"].transform("sum").fillna(0.0)
    df[AREA_COL] = 0.0

    for _, group in df.groupby(GROUP_COLS, sort=False):
        group_idx = list(group.index)
        group_total_area = float(group[TOTAL_AREA_COL].iloc[0]) if pd.notna(group[TOTAL_AREA_COL].iloc[0]) else 0.0
        matched = group[group["_matched"] & group["_plot_area_mu"].fillna(0).gt(0)]
        total_plot_area = float(matched["_plot_area_mu"].sum())

        if group_total_area <= 0:
            for idx in group_idx:
                issues.append(
                    {
                        "图斑编号": df.at[idx, ID_COL],
                        "所在镇村（镇街+村）": df.at[idx, GROUP_COLS[0]],
                        "养殖主体姓名（名称）": df.at[idx, GROUP_COLS[1]],
                        "问题描述": "养殖户总面积（亩）为空或无效，塘口面积已置为0",
                    }
                )
            continue

        if total_plot_area <= 0:
            for idx in group_idx:
                issues.append(
                    {
                        "图斑编号": df.at[idx, ID_COL],
                        "所在镇村（镇街+村）": df.at[idx, GROUP_COLS[0]],
                        "养殖主体姓名（名称）": df.at[idx, GROUP_COLS[1]],
                        "问题描述": "图斑编号未匹配到有效图斑面积，塘口面积已置为0",
                    }
                )
            continue

        allocations: dict[int, float] = {}
        for idx, row in matched.iterrows():
            allocations[idx] = round(group_total_area * float(row["_plot_area_mu"]) / total_plot_area, 2)

        diff = round(group_total_area - sum(allocations.values()), 2)
        if allocations and abs(diff) > 0:
            adjust_idx = matched["_plot_area_mu"].idxmax()
            allocations[adjust_idx] = round(allocations[adjust_idx] + diff, 2)

        for idx in group_idx:
            df.at[idx, AREA_COL] = allocations.get(idx, 0.0)

    return df, pd.DataFrame(issues)


def write_updated_excel(source_path: Path, output_path: Path, updated_df: pd.DataFrame, issue_df: pd.DataFrame) -> None:
    wb = load_workbook(source_path)
    ws = wb[wb.sheetnames[0]]
    header_row = detect_header_row(source_path) + 1
    data_start_row = header_row + 1

    headers = [normalize_text(ws.cell(row=header_row, column=i).value) for i in range(1, ws.max_column + 1)]
    header_map = {header: idx for idx, header in enumerate(headers, start=1) if header}

    for excel_row, (_, row) in enumerate(updated_df.iterrows(), start=data_start_row):
        ws.cell(row=excel_row, column=header_map[AREA_COL]).value = float(row[AREA_COL]) if pd.notna(row[AREA_COL]) else None
        ws.cell(row=excel_row, column=header_map[TOTAL_AREA_COL]).value = float(row[TOTAL_AREA_COL]) if pd.notna(row[TOTAL_AREA_COL]) else None
        ws.cell(row=excel_row, column=header_map[COUNT_COL]).value = int(row[COUNT_COL]) if pd.notna(row[COUNT_COL]) else None

    if "校验明细" in wb.sheetnames:
        del wb["校验明细"]
    issue_ws = wb.create_sheet("校验明细")
    if issue_df.empty:
        issue_ws.append(["图斑编号", "所在镇村（镇街+村）", "养殖主体姓名（名称）", "问题描述"])
        issue_ws.append(["", "", "", "未发现问题"])
    else:
        issue_ws.append(list(issue_df.columns))
        for _, issue in issue_df.iterrows():
            issue_ws.append(issue.tolist())

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()


def main() -> None:
    excel_path = Path(r"E:\全省养殖池溏上图入库普查\PY\宜兴市\20251112\0319\0509两鱼信息表_更新.xlsx")
    pond_path = Path(r"E:\全省养殖池溏上图入库普查\养殖信息统计\20250902\宜兴市池塘图斑库.gpkg")
    output_path = Path(r"E:\全省养殖池溏上图入库普查\PY\宜兴市\20251112\0319\0509两鱼信息表_更新_复核.xlsx")

    excel_df = load_excel_df(excel_path)
    pond_gdf = gpd.read_file(pond_path)
    updated_df, issue_df = build_updated_dataframe(excel_df, pond_gdf)
    write_updated_excel(excel_path, output_path, updated_df, issue_df)

    print(f"输入文件: {excel_path}")
    print(f"图斑库: {pond_path}")
    print(f"输出文件: {output_path}")
    print(f"记录数: {len(updated_df)}")
    print(f"问题数: {len(issue_df)}")


if __name__ == "__main__":
    main()
