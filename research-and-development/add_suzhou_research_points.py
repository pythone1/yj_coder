from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


PARENT = Path(r"D:/Users/Documents/WXWork/1688858186325806/Cache/File/2026-04")


def main():
    src = next(p for p in PARENT.glob("*_filled.xlsx") if not p.name.startswith("~$"))
    out = src.with_name(src.stem + "_含补充调研要点.xlsx")
    wb = load_workbook(src)
    ws = wb.worksheets[0]

    research_col = ws.max_column + 1
    ws.cell(1, research_col).value = "补充调研要点"
    header = ws.cell(1, research_col)
    header.fill = PatternFill("solid", fgColor="F4CCCC")
    header.font = Font(bold=True)
    header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    points = {
        2: "需确认：1）错装、漏装、反装、缺件分别有哪些典型场景；2）是否有现场图片、历史缺陷图片和人工判定记录；3）当前是否已有相机、光源、工控机、剔除机构等设备；4）检测节拍、拍摄位置、遮挡和反光情况；5）哪些缺陷必须零漏检，哪些可人工复核。",
        3: "需确认：1）标准操作步骤是否已有视频或作业指导书；2）工程师能否拆解每一步的动作、工具、零件和时间要求；3）现场摄像头角度是否能看清手部动作；4）漏步骤后希望如何提醒；5）98%-99%准确率是按单个动作、整套步骤还是工位结果统计。",
        4: "需确认：1）MES和设备数据目前能开放哪些字段；2）设备是否有温度、振动、电流、压力、报警等连续数据；3）历史故障、停机、维修工单是否完整；4）重点设备和停机损失排序；5）维修知识库资料来源、责任人和更新机制。",
        5: "需确认：1）当前在线检测系统的误检、漏检、无法判断样本是否可导出；2）准确率当前口径和目标口径；3）误检漏检主要发生在哪些产品、班次、光照和工位；4）是否允许调整相机、光源和拍摄位置；5）现场验收周期和连续运行要求。",
        6: "需确认：1）芯片良品分析涉及哪些数据表和系统；2）是否有晶圆图、缺陷图、量测数据、工艺参数和历史分析报告；3）经验库需要覆盖哪些典型问题；4）辅助报告的模板和审批流程；5）哪些结论必须工程师确认后才能输出。",
        7: "需确认：1）产品质量检测具体检测哪些缺陷和质量指标；2）是否有缺陷图片、质检记录、工艺参数和处置结果；3）现有质检流程中哪些环节最耗人；4）质量标准是否结构化；5）辅助建议输出后由谁确认、如何回写结果。",
        8: "需确认：1）希望先监测哪些生产关键步骤；2）每个步骤能采集哪些图片、视频和设备参数；3）是否有统一批次号和工序流转记录；4）过程异常如何触发提醒和工单；5）最终结果检测与过程检测如何关联追溯。",
    }

    for row, value in points.items():
        cell = ws.cell(row, research_col)
        cell.value = value
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 95

    ws.column_dimensions[ws.cell(1, research_col).column_letter].width = 62
    wb.save(out)
    print(out)


if __name__ == "__main__":
    main()
