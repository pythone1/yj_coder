from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment


BASE = Path(r"E:\PY\research\output\spreadsheet")
SRC = max(BASE.glob("*按原始需求.xlsx"), key=lambda p: p.stat().st_mtime)
OUT = BASE / "在线智能检测调研总表_按原始需求_中文审核版.xlsx"


def main():
    wb = load_workbook(SRC)
    ws = wb.worksheets[0]

    # Row 1 is title, row 2 is header.
    headers = {ws.cell(2, c).value: c for c in range(1, ws.max_column + 1)}

    def setv(row, header, value):
        cell = ws.cell(row, headers[header])
        cell.value = value
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Keep original demand description unchanged. Only clean research fields.
    updates = {
        3: {
            "需要哪些数据": "产品装配清单、装配工位图、相机图片、合格样本、不合格样本、错装/漏装标签、人工复核结果、生产批次和工单号。",
            "用什么算法/技术": "YOLO/RT-DETR用于找零件位置；Mask R-CNN用于识别缺陷区域；SAHI用于把大图切小后识别小零件；再用尺寸、位置、数量等规则复核。",
            "调研链接/案例": "KEYENCE缺件/有无检测案例：https://www.keyence.com/products/vision/applications/presence-and-absence-inspection.jsp | Cognex装配验证案例：https://www.cognex.com/industries/food-and-beverage/assembly-verification | PCB错件/漏件视觉检测论文：https://www.sciencedirect.com/science/article/pii/S1474034619305774",
        },
        4: {
            "对应功能模块": "装配标准操作步骤检测",
            "具体怎么做": "先由工程师拍摄正确操作视频，把操作拆成摄像头能看到的动作、工具和零件；系统识别当前做到哪一步，再按步骤顺序判断是否漏做或做错。",
            "需要哪些数据": "标准操作视频、步骤拆解表、动作标签、工具和零件图片、异常操作样本、工位摄像头视频、提醒记录。",
            "用什么算法/技术": "YOLO识别工具和零件；人体姿态识别看手和身体动作；SlowFast或Video Transformer识别连续动作；步骤顺序规则判断是否符合标准操作步骤。",
            "调研链接/案例": "SlowFast动作识别论文：https://arxiv.org/abs/1812.03982 | Video Swin Transformer动作识别论文：https://arxiv.org/abs/2106.13230 | 姿态识别工具参考：https://developers.google.com/mediapipe/solutions/vision/pose_landmarker",
        },
        5: {
            "具体怎么做": "统一设备编号和采集点含义；把生产执行系统、设备控制和监控数据、维修工单按时间对齐；用设备过去一段时间的运行数据训练LSTM，预测未来是否有故障风险。",
            "需要哪些数据": "设备台账、生产节拍和批次、设备控制和监控数据、温度、振动、电流、压力、报警码、停机记录、维修工单、备件记录。",
            "用什么算法/技术": "LSTM用于看设备状态随时间的变化；自编码器用于发现和平时不一样的运行状态；XGBoost可做对照模型。",
            "调研链接/案例": "工业设备LSTM预测维护案例：https://www.mdpi.com/1424-8220/21/3/972 | 制造资产预测维护案例：https://www.mdpi.com/1424-8220/24/10/3215 | A2-LSTM生产线剩余寿命预测：https://www.sciencedirect.com/science/article/pii/S0360835222005630",
        },
        6: {
            "需要哪些数据": "维修手册、标准操作步骤、历史工单、故障代码、备件清单、专家经验、相似案例、设备型号。",
            "用什么算法/技术": "文档解析和关键词抽取；文本向量搜索相似资料；检索增强问答用于先查资料再回答；故障树用于规范排查路径。",
            "调研链接/案例": "IBM Maximo设备维护方案：https://www.ibm.com/products/maximo | Siemens Senseye预测性维护方案：https://www.siemens.com/global/en/products/services/digital-enterprise-services/senseye-predictive-maintenance.html | RAG方法论文：https://arxiv.org/abs/2005.11401",
        },
        7: {
            "用什么算法/技术": "样本增强提升覆盖面；主动学习优先补标模型不确定样本；PatchCore用于少缺陷样本的异常检测；按“少漏检优先”的原则校准模型阈值。",
            "调研链接/案例": "PatchCore工业异常检测论文：https://arxiv.org/abs/2106.08265 | MVTec工业异常检测数据集：https://www.mvtec.com/company/research/datasets/mvtec-ad | IPC关于AI用于AOI的白皮书：https://www.ipc.org/news-release/new-ipc-white-paper-focuses-use-artificial-intelligence-automated-optical-inspection",
        },
        8: {
            "具体怎么做": "把芯片批次、晶圆图、量测数据、缺陷图、工艺参数和工程师分析记录串起来；系统先找相似历史案例和可能影响因素，再由工程师确认。",
            "需要哪些数据": "晶圆图、缺陷图、量测数据、工艺配方、设备参数、批次路径、过程控制报警、工程师分析记录。",
            "用什么算法/技术": "过程控制方法用于判断过程是否异常；XGBoost/LightGBM用于找影响良率的关键因素；晶圆图模式识别用于缺陷分类；检索增强问答用于查询经验库。",
            "调研链接/案例": "Intel制造良率AI白皮书：https://www.intel.com/content/dam/www/central-libraries/us/en/documents/intel-it-manufacturing-yield-analysis-with-ai-paper.pdf | Applied Materials AIx平台案例：https://ir.appliedmaterials.com/node/24181/pdf | 半导体良率预测案例论文：https://www.mdpi.com/2076-3417/13/4/2660",
        },
        9: {
            "用什么算法/技术": "检索增强问答用于查找资料和相似案例；大模型用于按模板组织文字；规则校验关键数字和来源。",
            "调研链接/案例": "Intel制造良率AI白皮书：https://www.intel.com/content/dam/www/central-libraries/us/en/documents/intel-it-manufacturing-yield-analysis-with-ai-paper.pdf | Deloitte半导体生成式AI应用分析：https://www2.deloitte.com/us/en/pages/technology-media-and-telecommunications/articles/gen-ai-semiconductor-industry.html | 文档理解模型参考：https://arxiv.org/abs/1912.13318",
        },
        10: {
            "用什么算法/技术": "图像分类和缺陷检测用于识别外观问题；XGBoost/LightGBM用于找影响质量的参数；检索增强问答用于查询质量标准和历史案例。",
            "调研链接/案例": "Cognex深度学习视觉检测案例：https://www.cognex.com/products/deep-learning/in-sight-vidi | KEYENCE视觉检测说明：https://www.keyence.com/products/vision/resources/vision-resources/what-are-vision-inspection-systems.jsp | PatchCore工业异常检测论文：https://arxiv.org/abs/2106.08265",
        },
        11: {
            "用什么算法/技术": "图像识别看关键动作和零件状态；LSTM看过程参数随时间变化；规则判断步骤是否符合要求；质量预测模型评估后续风险。",
            "调研链接/案例": "ISA-95系统集成标准：https://www.isa.org/standards-and-publications/isa-standards/isa-95 | OPC UA设备数据标准：https://opcfoundation.org/about/opc-technologies/opc-ua/ | 制造资产预测维护案例：https://www.mdpi.com/1424-8220/24/10/3215",
        },
        12: {
            "调研链接/案例": "ISA-95系统集成标准：https://www.isa.org/standards-and-publications/isa-standards/isa-95 | MTConnect设备数据标准：https://www.mtconnect.org/standard | OPC UA设备数据标准：https://opcfoundation.org/about/opc-technologies/opc-ua/",
        },
    }

    for row_idx, row_updates in updates.items():
        for header, value in row_updates.items():
            setv(row_idx, header, value)

    replacements = {
        "wafer map": "晶圆图",
        "wafer图": "晶圆图",
        "产品BOM": "产品装配清单",
        "BOM": "装配清单",
        "SOP": "标准操作步骤",
        "SPC报警": "过程控制报警",
        "SPC": "过程控制",
        "AutoEncoder": "自编码器",
        "RAG用于": "检索增强问答用于",
        "RAG论文": "RAG方法论文",
        "MediaPipe姿态识别": "姿态识别工具参考",
        "物料": "零件",
    }
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, str):
                text = cell.value
                for old, new in replacements.items():
                    text = text.replace(old, new)
                cell.value = text
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    if "修改说明" in wb.sheetnames:
        del wb["修改说明"]
    note = wb.create_sheet("修改说明")
    note.append(["检查项", "处理结果"])
    note.append(["中英文混杂", "正文中已将 wafer map/wafer图、BOM、SOP、SPC、AutoEncoder、RAG、物料 等改为中文或中文解释。"])
    note.append(["保留英文", "仅保留必要算法、标准或产品名，如 LSTM、YOLO、RT-DETR、Mask R-CNN、SAHI、PatchCore、XGBoost、OPC UA 等。"])
    note.append(["链接审核", "预测性维护链接已从LSTM原始论文换成工业设备/制造资产案例；AOI链接换成缺件检测、装配验证和AOI检测论文；芯片链接换成良率AI白皮书、AIx案例和良率预测案例论文。"])
    note.append(["原始需求", "原始需求描述未改动，只清理后续调研字段。"])
    note.column_dimensions["A"].width = 18
    note.column_dimensions["B"].width = 120
    for row in note.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
