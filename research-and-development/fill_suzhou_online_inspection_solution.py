from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


PARENT = Path(r"D:/Users/Documents/WXWork/1688858186325806/Cache/File/2026-04")


def find_source() -> Path:
    candidates = [
        f
        for f in PARENT.glob("*.xlsx")
        if not f.name.startswith("~$")
        and f.name.endswith("调研需求及应对方案梳理.xlsx")
        and "filled" not in f.name
    ]
    if candidates:
        return candidates[0]
    return [f for f in PARENT.glob("*.xlsx") if not f.name.startswith("~$") and f.stat().st_size == 13102][0]


def main():
    src = find_source()
    out = src.with_name(src.stem + "_filled.xlsx")

    wb = load_workbook(src)
    ws = wb.worksheets[0]
    solution_col = 5
    service_col = 6

    ws.cell(1, solution_col).value = "解决方案"
    ws.cell(1, service_col).value = "服务形式"
    for col in [solution_col, service_col]:
        cell = ws.cell(1, col)
        cell.fill = PatternFill("solid", fgColor="F4CCCC")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    fills = {
        2: (
            "建设装配视觉检测方案：固定工位相机、光源和拍摄位置，建立合格/不合格样本和错漏装判定规则；系统自动检查关键零件是否存在、位置是否正确，异常结果进入人工复核并回流样本库。",
            "现场工位勘察；相机/光源/拍摄方案建议；样本采集与标注；视觉模型训练和调优；边缘部署或对接现有检测系统；试运行、验收和人员培训。",
        ),
        3: (
            "建设标准操作步骤检测方案：先由工程师拍摄标准操作并拆解步骤，明确每一步要看的动作、工具和零件；系统识别当前步骤，发现漏做、错做或顺序异常时及时提示，并形成操作追溯记录。",
            "标准步骤梳理；工位视频采集与标注；动作识别模型验证；提示界面配置；试点上线；按关键步骤漏检率、误提醒率和提示延迟进行验收。",
        ),
        4: (
            "建设设备预测性维护和维修知识库方案：打通MES、设备监控数据和维修工单，使用LSTM等时序模型预测设备故障风险；同步整理故障现象、原因、处理措施和备件信息，形成可问答的维修知识库。",
            "设备点位和数据现状诊断；MES/设备数据/工单数据接入；预测模型开发；维修知识库整理；维修问答助手配置；与工单系统对接；上线后持续调优。",
        ),
        5: (
            "开展在线检测准确率提升专项：复盘现有误检、漏检和无法判断样本，区分是拍摄条件、样本不足、规则不清还是模型问题；针对性优化光源、样本库、识别模型和复核规则。",
            "现有系统诊断；误检漏检样本分析；补充采样和标注；模型重新训练与阈值调整；现场连续试运行；输出准确率提升报告和后续维护机制。",
        ),
        6: (
            "建设芯片良品分析和经验库方案：把批次、晶圆图、量测数据、缺陷图、工艺参数和工程师分析记录关联起来；系统查询相似历史案例，辅助分析可能原因，并生成分析报告草稿供工程师确认。",
            "数据字段梳理；历史报告和分析记录整理；良品/良率分析模型验证；经验库和问答功能配置；报告模板设计；小范围试点和工程师审核流程建设。",
        ),
        7: (
            "建设产品质量检测智能辅助方案：把质检记录、缺陷图片、工艺参数、批次信息和处置结果关联起来；系统辅助识别缺陷、查询质量标准和相似案例，给出质量风险提示和处理建议。",
            "质量标准梳理；质检数据和图片样本整理；缺陷识别模型验证；质量案例库建设；辅助判定界面配置；与现有质检流程对接；人员培训和试运行。",
        ),
        8: (
            "建设生产关键步骤智能监测方案：从最终结果检测扩展到关键工序过程检测，采集关键步骤图片/视频、设备参数和阶段质检结果；系统提前发现过程异常，并把过程结果与最终质检结果关联，形成全流程追溯。",
            "关键工序梳理；批次和工序编号规则设计；过程数据采集方案；视觉识别和过程预警模型验证；异常工单闭环配置；先选1-2个关键工序试点，再逐步推广。",
        ),
    }

    for row, (solution, service) in fills.items():
        ws.cell(row, solution_col).value = solution
        ws.cell(row, service_col).value = service
        ws.cell(row, solution_col).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, service_col).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 78

    ws.column_dimensions["E"].width = 48
    ws.column_dimensions["F"].width = 48
    wb.save(out)
    print(out)


if __name__ == "__main__":
    main()
