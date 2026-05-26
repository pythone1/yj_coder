from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment


PARENT = Path(r"D:/Users/Documents/WXWork/1688858186325806/Cache/File/2026-04")


def main():
    src = next(p for p in PARENT.glob("*最终版.xlsx") if not p.name.startswith("~$") and p.stat().st_size == 14689)
    out = src.with_name(src.stem.replace("_最终版", "_去模板版") + ".xlsx")
    wb = load_workbook(src)
    ws = wb.worksheets[0]

    solution_col = 5
    solutions = {
        2: "先从装配完成后的关键零件检查切入，把容易错装、漏装的部位梳理清楚，再确定相机、光源和拍摄位置。系统通过稳定拍照和视觉识别判断零件是否缺失、位置是否正确、数量是否一致，异常结果交给人工复核。复核结果继续沉淀为样本，用于后续扩大检测范围和减少误判。",
        3: "标准操作步骤检测需要先把人的操作变成可识别的步骤。由工程师确认标准流程，把拿取零件、安装、工具使用、确认动作等关键环节拆出来，再结合工位视频判断当前操作进度。系统发现漏步骤、顺序错误或超时后进行提醒，同时保留异常片段，便于后续追溯和培训。",
        4: "设备预测性维护建议与维修知识库一起建设。先选择停机影响较大的重点设备，接入设备运行数据、MES数据和维修工单，判断设备健康状态和故障风险。维修手册、历史故障、处理措施和备件信息同步整理成知识库，方便维修人员在预警后快速查询原因和处理办法。",
        5: "在线检测准确率提升应先做问题复盘。把现有系统中的误检、漏检、无法判断样本按产品、工位、光照、班次和缺陷类型分类，找出主要误差来源。现场条件能解决的优先调整相机、光源和拍摄位置，样本不足的补充样本，规则不清的明确判定口径，再进行模型和阈值优化。",
        6: "芯片良品分析的重点是把分散的数据和经验串起来。批次、晶圆图、缺陷图、量测数据、工艺参数和工程师分析记录需要形成关联，系统据此检索相似历史案例，给出可能影响因素和分析线索。大模型主要用于经验库问答和报告草稿整理，最终结论仍由工程师确认。",
        7: "产品质量检测智能辅助可以从质检人员最耗时、最容易判断不一致的环节入手。系统关联质检记录、缺陷图片、工艺参数、批次信息和处置结果，辅助识别缺陷类型，查询质量标准和相似案例，给出处理建议。质量人员确认后，结果回写为后续案例。",
        8: "生产关键步骤监测适合先选少量关键工序试点。对影响最终质量的步骤采集图片、视频、设备参数和阶段质检结果，识别过程中的异常并提前提醒。后续把过程记录与最终质检结果关联起来，形成从过程到结果的质量追溯链路，再逐步扩展到更多工序。",
    }

    for row, text in solutions.items():
        cell = ws.cell(row, solution_col)
        cell.value = text
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 105

    wb.save(out)
    print(out)


if __name__ == "__main__":
    main()
